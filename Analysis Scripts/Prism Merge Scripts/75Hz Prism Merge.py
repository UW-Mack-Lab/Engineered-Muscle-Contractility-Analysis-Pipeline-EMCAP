# -*- coding: utf-8 -*-
"""
Created on Wed Nov 19 23:22:10 2025

@author: phili
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jun 23 2025

@author: pb92
@version: 1.1.3

@description: A script to merge 75Hz relaxation analysis output files. It creates
             Prism-friendly tables for relaxation metrics plotted against days.

CHANGELOG:
v1.1.3 - ADAPTATION: Updated target metrics to match 75Hz Script v1.7.0.
       - Looks for 'Kinetics Peak-BK (μN)' and 'Max tetanic-BK (μN)'.
       - BACKWARD COMPATIBILITY: Falls back to 'BK corrected force (μN)' if new
         columns are missing (supports older files).
v1.1.2 - Removed Timecourse Max bar chart sheet.
"""

import os
import sys
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import uuid
from tqdm import tqdm

# --- File Selection & UI Functions ---
def select_files(root, title, file_types):
    """Opens a dialog to select multiple files."""
    return filedialog.askopenfilenames(parent=root, title=title, filetypes=file_types)

def select_save_path(root, title, file_types, default_name):
    """Opens a dialog to select a save location and filename."""
    return filedialog.asksaveasfilename(
        parent=root, title=title, filetypes=file_types, defaultextension=".xlsx", initialfile=default_name
    )

# --- Data Loading and Parsing Functions ---
def parse_day_from_filename(filename):
    """Extracts the 'Day' (e.g., 'd7') from the filename."""
    parts = os.path.basename(filename).split('_')
    day_part = next((part for part in parts if part.lower().startswith('d') and part[1:].isdigit()), None)
    return day_part

def load_and_combine_relaxation_data(root, input_files):
    """Loads and combines data from 75Hz relaxation analysis output files."""
    all_relaxation_data = []
    print("Loading and combining data from input files...")
    
    for file in tqdm(input_files, desc="Processing Files"):
        try:
            day = parse_day_from_filename(file)
            print(f"\n-> Processing file: {os.path.basename(file)}")
            if not day:
                print(f"   - WARNING: Could not determine Day from filename. Skipping.")
                continue
            print(f"   - Detected Day: {day}")

            xls = pd.ExcelFile(file)
            
            sheet_names_lower = [s.lower() for s in xls.sheet_names]
            target_sheet_lower = 'relaxation summary'

            if target_sheet_lower in sheet_names_lower:
                original_sheet_name = xls.sheet_names[sheet_names_lower.index(target_sheet_lower)]
                
                relaxation_df = pd.read_excel(xls, sheet_name=original_sheet_name)
                relaxation_df.columns = relaxation_df.columns.str.strip()
                relaxation_df['Day'] = day
                
                if all(col in relaxation_df.columns for col in ['Condition', 'Plate', 'Well']):
                    relaxation_df['Replicate_ID'] = (relaxation_df['Condition'].astype(str) + '_' + 
                                                     relaxation_df['Plate'].astype(str) + '_' + 
                                                     relaxation_df['Well'].astype(str))
                    all_relaxation_data.append(relaxation_df)
                    print(f"   - Loaded '{original_sheet_name}' sheet ({len(relaxation_df)} rows).")
                else:
                    print(f"   - WARNING: '{original_sheet_name}' in {os.path.basename(file)} is missing required columns (Condition, Plate, Well).")
            else:
                print(f"   - WARNING: '{target_sheet_lower}' sheet not found in {os.path.basename(file)}.")
                print(f"   - Available sheets: {', '.join(xls.sheet_names)}")

        except Exception as e:
            print(f"\nCould not read or process file {os.path.basename(file)}. Error: {e}")

    relaxation_df_combined = pd.concat(all_relaxation_data, ignore_index=True) if all_relaxation_data else pd.DataFrame()
    
    if not relaxation_df_combined.empty: 
        print(f"\nSuccessfully combined 'relaxation summary' data for {len(relaxation_df_combined['Replicate_ID'].unique())} replicates.")
    
    return relaxation_df_combined

# --- Excel Writing Function ---
def write_df_to_sheet(wb, sheet_name, df, add_summary_rows=False):
    """
    Writes a DataFrame to a new sheet as a formatted Excel table.
    Optionally adds summary rows for the max value and day of max.
    """
    safe_name = sheet_name.replace('%', 'Pct').replace('/', '_').replace(' ', '_').replace('(', '').replace(')', '')[:31]
    ws = wb.create_sheet(title=safe_name)
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    if df.shape[0] > 0:
        table_name = f"Table_{uuid.uuid4().hex[:10]}"
        table_ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
        try:
            tab = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, 
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)
        except ValueError:
            pass # Skip table if empty
    
    if add_summary_rows and df.shape[0] > 0:
        print(f"     - Adding Timecourse Max summary rows to '{safe_name}'...")
        summary_df = df.set_index('Day')
        
        # Calculate max and day of max
        max_values = summary_df.max()
        day_of_max = summary_df.idxmax()

        summary_start_row = ws.max_row + 2
        
        ws.cell(row=summary_start_row, column=1, value="Timecourse Max")
        for i, max_val in enumerate(max_values, 2):
            ws.cell(row=summary_start_row, column=i, value=max_val)

        ws.cell(row=summary_start_row + 1, column=1, value="Day of Max")
        for i, day_val in enumerate(day_of_max, 2):
            ws.cell(row=summary_start_row + 1, column=i, value=day_val)
            
    print(f"  - Sheet '{safe_name}' created.")


# --- Main File Creation Logic ---
def create_relaxation_prism_file(relaxation_df, output_path):
    """Creates the final merged Excel file with relaxation metric sheets."""
    if relaxation_df.empty:
        print("No relaxation data was loaded. Cannot create file.")
        return
        
    wb = Workbook()
    wb.remove(wb.active)
    print("\nCreating merged relaxation analysis file...")

    # --- METRIC DETECTION LOGIC (v1.1.3) ---
    
    # 1. Define the ideal (new) metrics we want
    potential_metrics = [
        'Kinetics Peak-BK (μN)',
        'Max tetanic-BK (μN)',
        'R50 Time', 
        'R80 Time', 
        'R90 Time', 
        'TT50P Time', 
        'TTP90 Time'
    ]
    
    # 2. Check for legacy metrics (Backward Compatibility)
    if 'Kinetics Peak-BK (μN)' not in relaxation_df.columns and 'BK corrected force (μN)' in relaxation_df.columns:
        print("   - Legacy Format Detected: Using 'BK corrected force (μN)' instead of 'Kinetics Peak-BK'.")
        # Swap the name in our target list
        potential_metrics[0] = 'BK corrected force (μN)'
        
    # 3. Filter list to only include metrics that actually exist in this dataset
    available_metrics = [col for col in potential_metrics if col in relaxation_df.columns]
    
    if not available_metrics:
        print("\nFATAL ERROR: None of the target metrics were found in the data.", file=sys.stderr)
        print(f"Looking for: {', '.join(potential_metrics)}")
        print(f"Found columns: {', '.join(relaxation_df.columns)}")
        return

    print(f"   - Metrics found for processing: {', '.join(available_metrics)}")

    condition_order = sorted(relaxation_df['Condition'].unique())
    condition_map = {condition: i for i, condition in enumerate(condition_order)}
    def get_condition_rank(replicate_id):
        return condition_map.get(str(replicate_id).split('_')[0], len(condition_map))

    print("\nStep 2: Creating Tidy Relaxation Summary sheet...")
    relaxation_df_tidy = relaxation_df.copy()
    relaxation_df_tidy['Day'] = relaxation_df_tidy['Day'].str[1:].astype(int)
    write_df_to_sheet(wb, "Tidy_Relaxation_Summary", relaxation_df_tidy)

    print("\nStep 3: Creating metric-specific sheets (vs. Day)...")
    for metric in available_metrics:
        print(f"   - Processing {metric}...")
        
        metric_data = relaxation_df[['Day', 'Replicate_ID', metric]].copy()
        metric_data['Day'] = metric_data['Day'].str[1:].astype(int)
        
        pivot = metric_data.pivot_table(index='Day', columns='Replicate_ID', values=metric, aggfunc='mean')
        pivot = pivot[sorted(pivot.columns, key=get_condition_rank)].reset_index()
        
        # Simplify sheet name (remove special chars)
        safe_metric_name = metric.replace(' ', '_').replace('%', 'Pct').replace('(', '').replace(')', '').replace('μN', 'uN')
        
        write_df_to_sheet(wb, f"{safe_metric_name}_vs_Day", pivot, add_summary_rows=True)

    print("\nStep 4: Creating condition summary mean sheets...")
    for metric in available_metrics:
        print(f"   - Processing {metric} (Mean by Condition)...")
        summary_data = relaxation_df.groupby(['Day', 'Condition'])[metric].agg(['mean', 'count']).reset_index()
        summary_data['Day'] = summary_data['Day'].str[1:].astype(int)
        
        pivot = summary_data.pivot_table(index='Day', columns='Condition', values='mean')
        pivot = pivot[sorted(pivot.columns)].reset_index()
        
        safe_metric_name = metric.replace(' ', '_').replace('%', 'Pct').replace('(', '').replace(')', '').replace('μN', 'uN')
        write_df_to_sheet(wb, f"{safe_metric_name}_Mean_by_Condition", pivot, add_summary_rows=True)

    try:
        print(f"\nSaving final workbook to: {output_path}")
        wb.save(output_path)
        print("✅ Analysis complete!")
        print(f"\nSummary:")
        print(f"- Processed {len(relaxation_df['Replicate_ID'].unique())} replicates")
        print(f"- Across {len(relaxation_df['Day'].unique())} days")
        print(f"- Plotted {len(available_metrics)} metrics and created all summary sheets.")
    except Exception as e:
        print(f"\nError saving file. Please ensure it is not open elsewhere. Error: {e}")

def main():
    root = tk.Tk()
    root.withdraw()
    print("=" * 60 + f"\nRelaxation 75Hz Analysis to Prism Formatter v1.1.3\n" + "=" * 60)
    
    input_files = select_files(root, "Select 75Hz relaxation analysis files to merge", [("Excel files", "*.xlsx")])
    if not input_files: 
        print("No files selected. Exiting.")
        return
    
    relaxation_df = load_and_combine_relaxation_data(root, input_files)
    
    if not relaxation_df.empty:
        output_path = select_save_path(root, "Select save location", [("Excel files", "*.xlsx")], "Relaxation_75Hz_Final_Tables.xlsx")
        if output_path:
            create_relaxation_prism_file(relaxation_df, output_path)
        else:
            print("No save location selected. Exiting.")
    else:
        print("\nNo data was loaded from the selected files. Nothing to save.")

if __name__ == "__main__":
    main()