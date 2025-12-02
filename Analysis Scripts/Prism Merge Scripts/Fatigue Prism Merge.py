# -*- coding: utf-8 -*-
"""
Created on Thu Nov 20 11:51:04 2025

@author: phili
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Nov 19 22:29:19 2025

@author: phili
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Jun 13 21:30:13 2025

@author: pb92
@version: 1.9.2

@description: A script to merge fatigue analysis output files. 
             It creates Prism-friendly tables for Fatigue Resistance Index (FRI),
             TimeAtXX metrics, and Force at specific stims.

CHANGELOG:
v1.9.2 - REMOVED: 'NormWaveform_vs_Time_dX' sheets are no longer generated to reduce
         confusion. Users should rely on individual daily files for waveform inspection.
v1.9.1 - ALIGNMENT: Updated to align with Fatigue Script v1.1.2 naming conventions.
       - LOGIC: Pivots 'Fatigue Resistance Index (%)' and 'FRI Block X (%)' directly
         from 'Fatigue Summary' sheet.
       - UI: Reordered tabs to prioritize FRI, TimeAt, and ForceAt metrics (Leftmost).
v1.9.0 - Initial pivoting logic for FRI.
"""

import os
import sys
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import uuid
from tqdm import tqdm

# --- File Selection & UI Functions ---
def select_files(root, title, file_types):
    """Opens a dialog to select multiple files."""
    return filedialog.askopenfilenames(parent=root, title=title, filetypes=file_types)

def select_save_path(root, title, file_types, default_name):
    """Opens a dialog to select a save location and filename."""
    return filedialog.asksaveasfilename(
        parent=root, title=title, filetypes=file_types, defaultextension=".xlsx", initialfile=default_name
    )

# --- Data Loading and Parsing Functions ---
def parse_day_from_filename(filename):
    """Extracts the 'Day' (e.g., 'd7') from the filename."""
    parts = os.path.basename(filename).split('_')
    day_part = next((part for part in parts if part.lower().startswith('d') and part[1:].isdigit()), None)
    return day_part

def load_and_combine_fatigue_data(root, input_files):
    """Loads and combines data from fatigue analysis output files (Detail and Summary only)."""
    all_details = []
    all_summaries = []
    
    print("Loading and combining data from input files...")
    
    for file in tqdm(input_files, desc="Processing Files"):
        try:
            day = parse_day_from_filename(file)
            print(f"\n-> Processing file: {os.path.basename(file)}")
            if not day:
                print(f"   - WARNING: Could not determine Day from filename. Skipping.")
                continue

            xls = pd.ExcelFile(file)
            
            # 1. Load Fatigue Detail (for TimeAt calculations and Curves)
            if 'Fatigue Detail' in xls.sheet_names:
                detail_df = pd.read_excel(xls, sheet_name='Fatigue Detail')
                detail_df.columns = detail_df.columns.str.strip()
                detail_df['Day'] = day
                if all(col in detail_df.columns for col in ['Condition', 'Plate', 'Well']):
                    detail_df['Replicate_ID'] = detail_df['Condition'].astype(str) + '_' + detail_df['Plate'].astype(str) + '_' + detail_df['Well'].astype(str)
                    all_details.append(detail_df)
            else:
                print(f"   - WARNING: 'Fatigue Detail' sheet not found.")

            # 2. Load Fatigue Summary (for Pre-calculated FRI metrics)
            if 'Fatigue Summary' in xls.sheet_names:
                summary_df = pd.read_excel(xls, sheet_name='Fatigue Summary')
                summary_df.columns = summary_df.columns.str.strip()
                summary_df['Day'] = day
                if all(col in summary_df.columns for col in ['Condition', 'Plate', 'Well']):
                    summary_df['Replicate_ID'] = summary_df['Condition'].astype(str) + '_' + summary_df['Plate'].astype(str) + '_' + summary_df['Well'].astype(str)
                    all_summaries.append(summary_df)
            else:
                print(f"   - WARNING: 'Fatigue Summary' sheet not found (FRI metrics may be missing).")
                    
        except Exception as e:
            print(f"\nCould not read or process file {os.path.basename(file)}. Error: {e}")

    details_df = pd.concat(all_details, ignore_index=True) if all_details else pd.DataFrame()
    summaries_df = pd.concat(all_summaries, ignore_index=True) if all_summaries else pd.DataFrame()

    return details_df, summaries_df

# --- Data Processing Functions ---
def calculate_time_at_force_percent(group, target_percent):
    """Calculates the time at which force drops to a target percentage using linear interpolation."""
    if 'Stim Peak Time (s)' not in group.columns:
        return np.nan
    group = group.sort_values('Stim Peak Time (s)')
    s = group.set_index('Stim Peak Time (s)')['Normalized Force (% of T1)']
    if s.empty or (s >= target_percent).all(): return np.nan
    if s.iloc[0] < target_percent: return s.index[0]
    try:
        above = s[s >= target_percent]
        below = s[s < target_percent]
        p_a_time, p_a_force = above.index[-1], above.iloc[-1]
        p_b_time, p_b_force = below.index[0], below.iloc[0]
        return p_a_time + ((target_percent - p_a_force) * (p_b_time - p_a_time)) / (p_b_force - p_a_force)
    except (IndexError, ZeroDivisionError):
        return np.nan

# --- Excel Writing Function ---
def write_df_to_sheet(wb, sheet_name, df):
    """Writes a DataFrame to a new sheet as a formatted Excel table."""
    safe_name = sheet_name.replace('%', 'Pct').replace('/', '_').replace(' ','_')[:31]
    ws = wb.create_sheet(title=safe_name)
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    table_name = f"Table_{uuid.uuid4().hex[:10]}"
    table_ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
    try:
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except ValueError:
        pass # Skip table creation if empty or invalid
    print(f"  - Sheet '{safe_name}' created.")

# --- Main File Creation Logic ---
def create_fatigue_prism_file(details_df, summaries_df, output_path):
    """Creates the final merged Excel file with reordered priority sheets."""
    if details_df.empty and summaries_df.empty:
        print("No data was loaded. Cannot create file.")
        return
        
    wb = Workbook()
    wb.remove(wb.active)
    print("\nCreating merged fatigue analysis file...")

    # Prepare Condition Mapping for Sorting Columns
    if not details_df.empty:
        condition_order = sorted(details_df['Condition'].unique())
        condition_map = {condition: i for i, condition in enumerate(condition_order)}
        def get_condition_rank(replicate_id):
            return condition_map.get(str(replicate_id).split('_')[0], len(condition_map))
    elif not summaries_df.empty:
        condition_order = sorted(summaries_df['Condition'].unique())
        condition_map = {condition: i for i, condition in enumerate(condition_order)}
        def get_condition_rank(replicate_id):
            return condition_map.get(str(replicate_id).split('_')[0], len(condition_map))

    # ==========================================
    # SECTION 1: PRIORITY METRICS (Shifted Left)
    # ==========================================
    
    # 1. Fatigue Resistance Index (Total) - Pivoted from Summary
    if not summaries_df.empty and 'Fatigue Resistance Index (%)' in summaries_df.columns:
        print("\nStep 1: Pivoting Total Fatigue Resistance Index...")
        summ_tidy = summaries_df.copy()
        summ_tidy['Day'] = summ_tidy['Day'].str[1:].astype(int)
        pivot = summ_tidy.pivot_table(index='Day', columns='Replicate_ID', values='Fatigue Resistance Index (%)')
        pivot = pivot[sorted(pivot.columns, key=get_condition_rank)].reset_index()
        write_df_to_sheet(wb, "Fatigue_Resistance_Total", pivot)

    # 2. Fatigue Resistance Per Block - Pivoted from Summary
    # Look for columns like "FRI Block 1 (%)"
    if not summaries_df.empty:
        print("\nStep 2: Pivoting Fatigue Resistance Per Block...")
        # Find columns containing "FRI Block"
        block_cols = [c for c in summaries_df.columns if 'FRI Block' in c]
        
        if block_cols:
            summ_tidy = summaries_df.copy()
            summ_tidy['Day'] = summ_tidy['Day'].str[1:].astype(int)
            # Melt block columns to pivot them
            melted_blocks = summ_tidy.melt(
                id_vars=['Day', 'Replicate_ID'], 
                value_vars=block_cols, 
                var_name='Block_Metric', 
                value_name='FRI_Value'
            )
            # Simplify Block Name for clean headers (Extract number from "FRI Block 1 (%)")
            melted_blocks['Block_Num'] = melted_blocks['Block_Metric'].str.extract(r'(\d+)').astype(int)
            
            pivot = melted_blocks.pivot_table(index=['Day', 'Replicate_ID'], columns='Block_Num', values='FRI_Value').reset_index()
            pivot.rename(columns={i: f'FRI_Block_{i}' for i in range(1, 5)}, inplace=True)
            write_df_to_sheet(wb, "Fatigue_Resistance_Per_Block", pivot)

    # 3. TimeAt Metrics (Calculated from Detail)
    if not details_df.empty:
        print("\nStep 3: Calculating TimeAt (Time to X% Force)...")
        details_df_calc = details_df.copy()
        details_df_calc['Day'] = details_df_calc['Day'].str[1:].astype(int)
        
        for percent in [75, 50]:
            try:
                results = details_df_calc.groupby(['Day', 'Replicate_ID']).apply(calculate_time_at_force_percent, target_percent=float(percent), include_groups=False).reset_index(name=f'TimeAt{percent}')
                pivot = results.pivot_table(index='Day', columns='Replicate_ID', values=f'TimeAt{percent}')
                pivot = pivot[sorted(pivot.columns, key=get_condition_rank)].reset_index()
                write_df_to_sheet(wb, f"TimeAt{percent}_Point", pivot)
            except Exception as e:
                print(f"  - Error calculating TimeAt{percent}: {e}")

    # 4. Force at Specific Stims (Pivoted from Detail)
    if not details_df.empty:
        print("\nStep 4: Extracting Force at Specific Twitches...")
        details_df_calc = details_df.copy()
        if 'Day' not in details_df_calc.columns or details_df_calc['Day'].dtype == object:
             details_df_calc['Day'] = details_df_calc['Day'].str[1:].astype(int)

        for stim in [5, 10, 20]:
            stim_df = details_df_calc[details_df_calc['Stim Number'] == stim].copy()
            if not stim_df.empty:
                pivot = stim_df.pivot_table(index='Day', columns='Replicate_ID', values='Normalized Force (% of T1)')
                pivot = pivot[sorted(pivot.columns, key=get_condition_rank)].reset_index()
                write_df_to_sheet(wb, f"Force_at_Stim_{stim}", pivot)

    # ==========================================
    # SECTION 2: CURVES & RAW DATA (Shifted Right)
    # ==========================================

    if not details_df.empty:
        print("\nStep 5: Creating Fatigue Curves (Day-Segregated)...")
        unique_days = sorted(details_df['Day'].unique(), key=lambda d: int(d[1:]))
        
        # NormForce vs Stim
        for day in unique_days:
            day_details = details_df[details_df['Day'] == day]
            pivot = day_details.pivot_table(index='Stim Number', columns='Replicate_ID', values='Normalized Force (% of T1)')
            pivot = pivot[sorted(pivot.columns, key=get_condition_rank)].reset_index()
            write_df_to_sheet(wb, f"NormForce_vs_Stim_d{day[1:]}", pivot)

        # NormForce vs Time
        for day in unique_days:
            day_details = details_df[details_df['Day'] == day]
            if 'Stim Peak Time (s)' in day_details.columns:
                pivot = day_details.pivot_table(index='Stim Peak Time (s)', columns='Replicate_ID', values='Normalized Force (% of T1)')
                pivot = pivot[sorted(pivot.columns, key=get_condition_rank)].reset_index()
                write_df_to_sheet(wb, f"NormForce_vs_Time_d{day[1:]}", pivot)

    if not details_df.empty:
        print("\nStep 6: Saving Tidy Data Backup...")
        details_df_tidy = details_df.copy()
        details_df_tidy['Day'] = details_df_tidy['Day'].str[1:].astype(int)
        write_df_to_sheet(wb, "Tidy_Fatigue_Detail", details_df_tidy)

    try:
        print(f"\nSaving final workbook to: {output_path}")
        wb.save(output_path)
        print("✅ Analysis complete!")
    except Exception as e:
        print(f"\nError saving file. Please ensure it is not open elsewhere. Error: {e}")

def main():
    root = tk.Tk()
    root.withdraw()
    print("=" * 60 + f"\nFatigue Analysis to Prism Formatter v1.9.2\n" + "=" * 60)
    input_files = select_files(root, "Select fatigue analysis files to merge", [("Excel files", "*.xlsx")])
    if not input_files: 
        print("No files selected. Exiting.")
        return
        
    details_df, summaries_df = load_and_combine_fatigue_data(root, input_files)
    
    if not details_df.empty or not summaries_df.empty:
        output_path = select_save_path(root, "Select save location", [("Excel files", "*.xlsx")], "Fatigue_Final_Filtered_Tables.xlsx")
        if output_path:
            create_fatigue_prism_file(details_df, summaries_df, output_path)
        else:
            print("No save location selected. Exiting.")
    else:
        print("\nNo data was loaded from the selected files. Nothing to save.")

if __name__ == "__main__":
    main()