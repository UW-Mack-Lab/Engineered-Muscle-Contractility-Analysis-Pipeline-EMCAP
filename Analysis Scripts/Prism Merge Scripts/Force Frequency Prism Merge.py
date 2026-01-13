
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 06 23:30:00 2025


@version: 17.3.0

@description: A script to merge FvF analysis files and format for GraphPad Prism.
             This definitive version creates simple, flat, filterable tables
             by flattening pivot tables and restoring table formatting.

CHANGELOG:
v17.3.0 - FEATURE: Removed 'd' prefix from Day values in all output sheets.
          The 'Day' column now contains numeric values for proper sorting in Excel.
v17.2.0 - FEATURE: Changed 'Absolute_Max_Force' to 'Max tetanic-bk force' and now
          uses 100Hz data for this calculation, per user request. This removes
          the now-redundant 'Hz_at_Abs_Max_Force' sheet.
v17.1.0 - PATCH: Fixed a TypeError caused by a missing argument in a call to
          the write_df_to_sheet function.
v17.0.0 - Final layout redesign to produce simple, flat tables.
"""

import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import uuid
from tqdm import tqdm

# --- File Selection Functions ---
def select_files(root, title, file_types):
    return filedialog.askopenfilenames(parent=root, title=title, filetypes=file_types)

def select_save_path(root, title, file_types, default_name):
    return filedialog.asksaveasfilename(
        parent=root, title=title, filetypes=file_types, defaultextension=".xlsx", initialfile=default_name
    )

# --- Data Loading and Parsing Functions ---
def parse_day_from_filename(filename):
    parts = os.path.basename(filename).split('_')
    day_part = next((part for part in parts if part.lower().startswith('d') and part[1:].isdigit()), None)
    return day_part

def load_and_combine_data(root, input_files):
    all_summaries = []
    print("Loading and combining data from input files...")
    for i, file in enumerate(input_files):
        try:
            day = parse_day_from_filename(file)
            print(f"\n-> Processing file: {os.path.basename(file)}")
            if not day:
                print(f"   - WARNING: Could not determine Day from filename. Grouping under 'Day_Unknown'.")
                day = "Day_Unknown"
            print(f"   - Detected Day: {day}")

            xls = pd.ExcelFile(file)
            if 'Relaxation Summary' in xls.sheet_names:
                summary_df = pd.read_excel(xls, sheet_name='Relaxation Summary')
                summary_df['Day'] = day
                summary_df['Condition'] = summary_df['Condition'].astype(str).str.strip()

                if 'Plate' in summary_df.columns:
                    print(f"   - 'Plate' column found. Automatically creating Replicate IDs.")
                    summary_df['Plate'] = summary_df['Plate'].astype(str)
                    summary_df['Replicate_ID'] = summary_df['Condition'] + '_' + summary_df['Plate'] + '_' + summary_df['Well']
                else:
                    print(f"   - WARNING: 'Plate' column not found in this file.")
                    plate_id = simpledialog.askstring("Input Required",f"'Plate' column is missing.\n\nEnter a unique ID for the plate in this file:\n\n{os.path.basename(file)}",
                        initialvalue=f"Plate{i+1}", parent=root)
                    if not plate_id:
                        print(f"   - Skipping file: {os.path.basename(file)}")
                        continue
                    summary_df['Replicate_ID'] = summary_df['Condition'] + '_' + plate_id + '_' + summary_df['Well']
                all_summaries.append(summary_df)
        except Exception as e:
            print(f"\nCould not read file {os.path.basename(file)}. Error: {e}")
            
    if not all_summaries:
        return None
    return pd.concat(all_summaries, ignore_index=True)

# --- Excel Writing Function ---
def write_df_to_sheet(wb, sheet_name, df):
    safe_name = sheet_name.replace(' (%)', '_Norm').replace(' (μN)', '_uN').replace(' ','_')[:31]
    ws = wb.create_sheet(title=safe_name)
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    table_name = f"Table_{uuid.uuid4().hex[:10]}"
    table_ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
    tab = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    print(f"  - Sheet '{safe_name}' created as a filterable table.")

# --- Main File Creation Logic ---
def create_prism_file(summary_df, output_path):
    if summary_df is None: return
    wb = Workbook()
    wb.remove(wb.active)
    print("\nCreating simple, filterable tables...")

    condition_order = summary_df['Condition'].unique()
    condition_map = {condition: i for i, condition in enumerate(condition_order)}
    def get_condition_rank(replicate_id):
        for condition, rank in condition_map.items():
            if str(replicate_id).startswith(condition): return rank
        return len(condition_map)

    print("\nStep 1: Creating Tidy Data Summary sheet...")
    tidy_summary_df = summary_df.copy()
    tidy_summary_df['Day'] = tidy_summary_df['Day'].apply(
        lambda d: int(d[1:]) if isinstance(d, str) and d.lower().startswith('d') and d[1:].isdigit() else d
    )
    write_df_to_sheet(wb, "Tidy_Data_Summary", tidy_summary_df)
    
    print("\nStep 2: Creating Day-segregated Force-Frequency sheets...")
    unique_days = sorted(summary_df['Day'].unique(), key=lambda d: int(d[1:]) if (isinstance(d, str) and d.lower().startswith('d')) else 999)
    freq_metrics = ['FF Normalized (%)', 'Max tetanic-BK (μN)']
    for day in unique_days:
        if not (isinstance(day, str) and day.lower().startswith('d')): continue # Skip 'Day_Unknown'
        
        day_num_str = day[1:]
        day_summary = summary_df[summary_df['Day'] == day].copy()
        print(f"  - Generating F-F curve sheets for Day {day_num_str}...")
        for metric in freq_metrics:
            if metric in day_summary.columns and not day_summary[metric].isnull().all():
                pivot_df = day_summary.pivot_table(index='Hz', columns='Replicate_ID', values=metric)
                sorted_columns = sorted(pivot_df.columns, key=get_condition_rank)
                pivot_df = pivot_df[sorted_columns]
                pivot_df = pivot_df.reset_index()
                write_df_to_sheet(wb, f"{metric}_vs_Freq_{day_num_str}", pivot_df)

    print("\nStep 3: Creating consolidated 100Hz summary sheets...")
    
    summary_100hz = summary_df[summary_df['Hz'] == 100].copy()
    if not summary_100hz.empty:
        summary_100hz['day_num'] = summary_100hz['Day'].str.extract(r'(\d+)').astype(int)
        summary_100hz = summary_100hz.sort_values('day_num')
        
        # Create 'Max tetanic-bk force' sheet using 100Hz data
        if 'Max tetanic-BK (μN)' in summary_100hz.columns:
            print("  - Generating 'Max tetanic-bk force' summary sheet (from 100Hz data)...")
            pivot_max_force = summary_100hz.pivot_table(index='Day', columns='Replicate_ID', values='Max tetanic-BK (μN)')
            sorted_columns_max = sorted(pivot_max_force.columns, key=get_condition_rank)
            pivot_max_force = pivot_max_force[sorted_columns_max]
            pivot_max_force = pivot_max_force.reset_index()
            pivot_max_force['Day'] = pivot_max_force['Day'].str[1:].astype(int) # Convert 'd7' to 7
            write_df_to_sheet(wb, 'Max tetanic-bk force', pivot_max_force)
        
        # Create Relaxation Kinetics sheets
        print("  - Generating 100Hz Relaxation Kinetics summary sheets...")
        kinetics_cols = ['R10 Time', 'R50 Time', 'R80 Time', 'R90 Time', 'TT50P Time', 'TTP90 Time']
        for metric in kinetics_cols:
            if metric in summary_100hz.columns and not summary_100hz[metric].isnull().all():
                pivot_kinetics = summary_100hz.pivot_table(index='Day', columns='Replicate_ID', values=metric)
                sorted_columns_kinetics = sorted(pivot_kinetics.columns, key=get_condition_rank)
                pivot_kinetics = pivot_kinetics[sorted_columns_kinetics]
                pivot_kinetics = pivot_kinetics.reset_index()
                pivot_kinetics['Day'] = pivot_kinetics['Day'].str[1:].astype(int) # Convert 'd7' to 7
                write_df_to_sheet(wb, f'{metric}_at_100Hz', pivot_kinetics)
    else:
        print("  - No 100Hz data found to create consolidated summary sheets.")
            
    try:
        print(f"\nSaving final workbook to: {output_path}")
        wb.save(output_path)
        print("✅ Analysis complete!")
    except Exception as e:
        print(f"\nError saving the file. Please ensure it is not open elsewhere. Error: {e}")

def main():
    root = tk.Tk()
    root.withdraw()
    print("=" * 60)
    print("FvF Analysis to Prism Formatter v17.3")
    print("=" * 60)
    input_files = select_files(root, "Select FvF analysis files to merge", [("Excel files", "*.xlsx")])
    if not input_files: return
    summary_df = load_and_combine_data(root, input_files)
    if summary_df is not None:
        output_path = select_save_path(root, "Select where to save the final file", [("Excel files", "*.xlsx")], "FvF_Final_Filtered_Tables.xlsx")
        if not output_path: return
        create_prism_file(summary_df, output_path)

if __name__ == "__main__":

    main()
