# -*- coding: utf-8 -*-
"""
Created on Fri Nov 21 11:25:18 2025

@author: phili
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jun 23 2025

@author: pb92 (adapted from Relaxation Prism Merge)

@version: 1.0.6

@description: A script to merge twitch analysis output files. It creates
             Prism-friendly tables for twitch contractile and relaxation metrics
             by combining data from both 'Contractile Averages' and 'Relaxation Detail'
             sheets.

CHANGELOG:
v1.0.6 - CLEANUP: Automatically removes redundant 'Day_contractile' and 'Day_relaxation'
         columns generated during the merge process.
v1.0.5 - UI: Reordered 'Tidy_Twitch_Master' columns to prioritize individual force.
"""

import os
import sys
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import uuid
from tqdm import tqdm

# --- File Selection & UI Functions ---
def select_files(root, title, file_types):
    """Opens a dialog to select multiple files."""
    return filedialog.askopenfilenames(parent=root, title=title, filetypes=file_types)

def select_save_path(root, title, file_types, default_name):
    """Opens a dialog to select a save location and filename."""
    return filedialog.asksaveasfilename(
        parent=root, title=title, filetypes=file_types, defaultextension=".xlsx", initialfile=default_name
    )

# --- Data Loading and Parsing Functions ---
def parse_day_from_filename(filename):
    """Extracts the 'Day' (e.g., 'd10') from the filename."""
    parts = os.path.basename(filename).split('_')
    day_part = next((part for part in parts if part.lower().startswith('d') and part[1:].isdigit()), None)
    return day_part

def load_and_combine_twitch_data(root, input_files):
    """Loads and combines data from Twitch analysis output files."""
    all_twitch_data = []
    print("Loading and combining data from input files...")
    
    for file in tqdm(input_files, desc="Processing Files"):
        try:
            day = parse_day_from_filename(file)
            print(f"\n-> Processing file: {os.path.basename(file)}")
            if not day:
                print(f"   - WARNING: Could not determine Day from filename. Skipping.")
                continue
            print(f"   - Detected Day: {day}")

            xls = pd.ExcelFile(file)
            
            sheet_names_lower = [s.lower() for s in xls.sheet_names]
            
            contractile_sheet_name = None
            relaxation_sheet_name = None
            
            if 'contractile averages' in sheet_names_lower:
                contractile_sheet_name = xls.sheet_names[sheet_names_lower.index('contractile averages')]
            else:
                print(f"   - WARNING: 'Contractile Averages' sheet not found in {os.path.basename(file)}.")

            if 'relaxation detail' in sheet_names_lower:
                relaxation_sheet_name = xls.sheet_names[sheet_names_lower.index('relaxation detail')]
            else:
                print(f"   - WARNING: 'Relaxation Detail' sheet not found in {os.path.basename(file)}.")

            if contractile_sheet_name and relaxation_sheet_name:
                contractile_df = pd.read_excel(xls, sheet_name=contractile_sheet_name)
                relaxation_df = pd.read_excel(xls, sheet_name=relaxation_sheet_name)
                
                contractile_df.columns = contractile_df.columns.str.strip()
                relaxation_df.columns = relaxation_df.columns.str.strip()

                merge_cols = ['Condition', 'Plate', 'Well']
                if not all(col in contractile_df.columns for col in merge_cols) or \
                   not all(col in relaxation_df.columns for col in merge_cols):
                    print(f"   - WARNING: Missing key columns for merging. Skipping.")
                    continue
                    
                merged_df = pd.merge(contractile_df, relaxation_df, on=merge_cols, how='inner', suffixes=('_contractile', '_relaxation'))
                
                # [FIX] Remove redundant Day columns created by the merge
                redundant_cols = ['Day_contractile', 'Day_relaxation']
                merged_df.drop(columns=[c for c in redundant_cols if c in merged_df.columns], inplace=True)

                merged_df['Replicate_ID'] = (merged_df['Condition'].astype(str) + '_' + 
                                             merged_df['Plate'].astype(str) + '_' + 
                                             merged_df['Well'].astype(str))
                merged_df['Day'] = day
                all_twitch_data.append(merged_df)
            else:
                print(f"   - Skipping file due to missing sheets.")

        except Exception as e:
            print(f"\nCould not read or process file {os.path.basename(file)}. Error: {e}")

    twitch_df_combined = pd.concat(all_twitch_data, ignore_index=True) if all_twitch_data else pd.DataFrame()
    
    if not twitch_df_combined.empty: 
        print(f"\nSuccessfully combined data for {len(twitch_df_combined['Replicate_ID'].unique())} total replicates.")
    
    return twitch_df_combined


# --- Excel Writing Function ---
def write_df_to_sheet(wb, sheet_name, df):
    """Writes a DataFrame to a new sheet as a formatted Excel table."""
    safe_name = sheet_name.replace('%', 'Pct').replace('/', '_').replace(' ', '_').replace('(', '').replace(')', '')[:31]
    ws = wb.create_sheet(title=safe_name)
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    if df.shape[0] > 0:
        table_name = f"Table_{uuid.uuid4().hex[:10]}"
        table_ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, 
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    
    print(f"  - Sheet '{safe_name}' created.")

# --- Main File Creation Logic ---
def create_twitch_prism_file(twitch_df, output_path):
    """Creates the final merged Excel file with twitch metric sheets."""
    if twitch_df.empty:
        print("No twitch data was loaded. Cannot create file.")
        return
        
    wb = Workbook()
    wb.remove(wb.active)
    print("\nCreating merged twitch analysis file...")

    print("\nStep 1: Verifying required columns...")
    
    target_metrics = [
        'Mean BK Corrected Force (μN)',
        'Max BK Corrected Force (μN)',
        'BK Force CV (%)', 
        'Early Twitch Mean (T1-T5) (μN)', 
        'Late Twitch Mean (Last 5) (μN)', 
        'Fatigue Index (Late/Early)',
        'R50 Time Avg (s)', 
        'R80 Time Avg (s)', 
        'R90 Time Avg (s)', 
        'TT50P Time Avg (s)', 
        'TTP90 Time Avg (s)'
    ]
    
    available_metrics = [col for col in target_metrics if col in twitch_df.columns]
    
    if not available_metrics:
        print("\nFATAL ERROR: None of the target metrics for plotting were found in the data.", file=sys.stderr)
        return

    print(f"   - Metrics found for plotting: {', '.join(available_metrics)}")

    condition_order = sorted(twitch_df['Condition'].unique())
    condition_map = {condition: i for i, condition in enumerate(condition_order)}
    def get_condition_rank(replicate_id):
        return condition_map.get(str(replicate_id).split('_')[0], len(condition_map))

    # --- Tidy Master Creation (With Column Reordering) ---
    print("\nStep 2: Creating Tidy Twitch Master sheet with all data...")
    twitch_df_tidy = twitch_df.copy()
    twitch_df_tidy['Day'] = twitch_df_tidy['Day'].str[1:].astype(int)
    
    # Define Preferred Column Order to reduce confusion
    preferred_order = [
        'Day', 'Replicate_ID', 'Condition', 'Plate', 'Well', 'Stim Number',
        'BK corrected force (μN)',        # <--- Individual Stim Force
        'Normalized Force (% of T1)',     # <--- Individual Normalization
        'Mean BK Corrected Force (μN)',   # <--- Well Average (Moved later)
        'Max BK Corrected Force (μN)',    # <--- Well Max (Moved later)
    ]
    
    # Reorder: Keep preferred columns first, append the rest found in dataframe
    existing_cols = [c for c in preferred_order if c in twitch_df_tidy.columns]
    remaining_cols = [c for c in twitch_df_tidy.columns if c not in existing_cols]
    final_col_order = existing_cols + remaining_cols
    
    twitch_df_tidy = twitch_df_tidy[final_col_order]
    write_df_to_sheet(wb, "Tidy_Twitch_Master", twitch_df_tidy)

    print("\nStep 3: Creating metric-specific sheets (vs. Day)...")
    for metric in available_metrics:
        print(f"   - Processing {metric}...")
        
        metric_data = twitch_df[['Day', 'Replicate_ID', metric]].copy()
        metric_data['Day'] = metric_data['Day'].str[1:].astype(int)
        
        pivot = metric_data.pivot_table(index='Day', columns='Replicate_ID', values=metric)
        pivot = pivot[sorted(pivot.columns, key=get_condition_rank)].reset_index()
        
        safe_metric_name = metric.replace(' ', '_').replace('%', 'Pct').replace('/', '_').replace('(', '').replace(')', '')
        write_df_to_sheet(wb, f"{safe_metric_name}_vs_Day", pivot)

    try:
        print(f"\nSaving final workbook to: {output_path}")
        wb.save(output_path)
        print("✅ Analysis complete!")
        print(f"\nSummary:")
        print(f"- Processed {len(twitch_df['Replicate_ID'].unique())} replicates")
        print(f"- Across {len(twitch_df['Day'].unique())} days")
        print(f"- Plotted {len(available_metrics)} metrics: {', '.join(available_metrics)}")
    except Exception as e:
        print(f"\nError saving file. Please ensure it is not open elsewhere. Error: {e}")

def main():
    root = tk.Tk()
    root.withdraw()
    print("=" * 60 + f"\nTwitch 1Hz Analysis to Prism Formatter v1.0.6\n" + "=" * 60)
    
    input_files = select_files(root, "Select Twitch analysis files to merge", [("Excel files", "*.xlsx")])
    if not input_files: 
        print("No files selected. Exiting.")
        return
    
    twitch_df = load_and_combine_twitch_data(root, input_files)
    
    if not twitch_df.empty:
        output_path = select_save_path(root, "Select save location", [("Excel files", "*.xlsx")], "Twitch_Final_Tables.xlsx")
        if output_path:
            create_twitch_prism_file(twitch_df, output_path)
        else:
            print("No save location selected. Exiting.")
    else:
        print("\nNo data was loaded from the selected files. Nothing to save.")

if __name__ == "__main__":
    main()