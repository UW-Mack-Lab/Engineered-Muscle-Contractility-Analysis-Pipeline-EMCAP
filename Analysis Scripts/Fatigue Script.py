# -*- coding: utf-8 -*-
"""
Created on Thu May 22 2025

@author: pb92
@version: 1.1.3 - CSV Support Added

DESCRIPTION:
Specialized script for analyzing muscle fatigue protocol data.
Protocol: 4 blocks of 10 stimulations with decreasing rest periods.

CHANGELOG:
v1.1.3 - ADDED: CSV support for Mantarray Controller files.
v1.1.2 - NAMING & FORMATTING: Standardized naming.
       - Total metric: "Fatigue Resistance Index (%)"
       - Block metric: "FRI Block X (%)"
       - Excel Formatting: explicitly catches "Index" and "FRI" to force 2 decimal rounding.
v1.1.1 - Formatting fix for FRI rounding.
v1.1.0 - Integrated Fatigue Resistance Index metrics.
"""

import os
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from tqdm import tqdm
import uuid
import json

# ==========================================
#        CONFIGURATION SECTION
# ==========================================

# Recording Settings
MAX_RECORDING_TIME = 65.0    # Maximum time (seconds) to process. Trims data beyond this point.

# Fatigue Protocol Definition (Fixed)
FATIGUE_PROTOCOL = {
    "initial_delay": 3.0,  # seconds
    "stim_duration": 0.5,  # ~ms total active duration
    "blocks": [
        {"iterations": 10, "rest_duration": 1.5},  # Block 1: 1.5s rest
        {"iterations": 10, "rest_duration": 1.0},  # Block 2: 1.0s rest  
        {"iterations": 10, "rest_duration": 0.5},  # Block 3: 0.5s rest
        {"iterations": 10, "rest_duration": 0.25}  # Block 4: 0.25s rest
    ]
}

# Analysis parameters
BASELINE_WINDOW = 0.2    # Window before T0 for baseline calculation
PEAK_SEARCH_WINDOW = 0.7 # Window after T0 to search for peak

# ==========================================
#      END CONFIGURATION SECTION
# ==========================================

T0_PAIRS = []            # Will store all stimulation start times
BLOCK_INFO = []          # Will store block information for each stim

def select_files(title, file_types):
    root = tk.Tk()
    root.withdraw()
    files = filedialog.askopenfilenames(title=title, filetypes=file_types)
    return files

def select_directory(title):
    root = tk.Tk()
    root.withdraw()
    directory = filedialog.askdirectory(title=title)
    return directory

def df_to_excel_sheet(wb, sheet_name, df, start_cell='A1'):
    # Format numbers with consistent decimal places before sending to Excel
    df_formatted = df.copy()
    
    # Apply formatting to numeric columns based on their content
    for col in df_formatted.columns:
        # 1. Time Columns -> 2 decimals
        if 'Time' in col and col != 'Time (seconds)' and df_formatted[col].dtype in ['float64', 'float32']:
            df_formatted[col] = df_formatted[col].round(2)
            
        # 2. Force, Threshold, StdDev, Index, FRI -> Specific rules
        elif ('Force' in col or 'Threshold' in col or 'StdDev' in col or 'Index' in col or 'FRI' in col) and df_formatted[col].dtype in ['float64', 'float32']:
            
            # PERCENTAGE-LIKE metrics -> 2 decimals
            # Includes: Normalized Force, Coefficients of Variation, FRI, Indices
            if ('Normalized Force' in col or 
                'Mean Normalized Trace' in col or 
                'Fatigue Resistance Index' in col or 
                'FRI' in col or 
                'CV' in col): 
                 df_formatted[col] = df_formatted[col].round(2)
            
            # RAW FORCE metrics -> 4 decimals (high precision)
            elif 'Force' in col or 'Threshold' in col : 
                 pass # Keep full precision (or round to 4 if preferred)
            
            else:
                 df_formatted[col] = df_formatted[col].round(4)
                 
        elif col == 'Relative Time (ms)':
            df_formatted[col] = df_formatted[col].round(0)
    
    ws = wb.create_sheet(title=sheet_name)
    
    rows = dataframe_to_rows(df_formatted, index=False, header=True)
    for r_idx, row_val in enumerate(rows, 1): 
        for c_idx, value in enumerate(row_val, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    max_row = len(df_formatted) + 1  # +1 for header
    max_col = len(df_formatted.columns)
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    
    safe_table_name = f"Table_{uuid.uuid4().hex[:8]}" 
    
    try:
        tab = Table(displayName=safe_table_name, ref=f"{start_cell}:{end_cell}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except ValueError as e:
        print(f"Warning: Could not create table in sheet '{sheet_name}': {str(e)}")
    
    max_sample_rows = 100 
    for column_cells in ws.columns: 
        max_length = 0
        column_letter = column_cells[0].column_letter 
        for cell in column_cells[:min(max_sample_rows, len(df_formatted)+1)]: 
            if cell.value:
                cell_value_str = str(cell.value)
                max_length = max(max_length, len(cell_value_str))
        adjusted_width = max(10, min(max_length + 2, 50)) 
        ws.column_dimensions[column_letter].width = adjusted_width

def parse_filename(filename):
    parts = filename.split('_')
    day = next((part for part in parts if part.lower().startswith('d') and part[1:].isdigit()), None)
    if day:
        try:
            day_index = parts.index(day)
            if day_index + 1 < len(parts) and day_index + 2 < len(parts):
                plate_name = parts[day_index + 1]
                protocol = parts[day_index + 2]
                return day, protocol, plate_name
        except ValueError: 
             pass 
    print(f"Warning: Could not parse day, protocol, and plate name from filename: {filename}")
    return None, None, None

def smooth_data(df, column, window_size=3):
    if column not in df.columns:
        print(f"Warning: Column '{column}' not found for smoothing. Returning original DataFrame.")
        return df
    df[f'{column}_Smoothed'] = df.groupby('Well')[column].rolling(
        window=window_size, center=True, min_periods=1).mean().reset_index(level=0, drop=True)
    return df

def calculate_fatigue_protocol_times(protocol_def, max_time): 
    """Calculate all T0 stimulation times for the fatigue protocol"""
    global T0_PAIRS, BLOCK_INFO
    
    T0_PAIRS = []
    BLOCK_INFO = []
    
    current_time = protocol_def["initial_delay"]
    stim_duration = protocol_def["stim_duration"]
    
    print(f"Calculating fatigue protocol timing:")
    print(f"  Initial delay: {protocol_def['initial_delay']}s")
    print(f"  Stimulation duration: {stim_duration}s")
    
    total_stim_count = 0 
    for block_num, block_params in enumerate(protocol_def["blocks"], 1): 
        iterations = block_params["iterations"]
        rest_duration = block_params["rest_duration"]
        
        print(f"  Block {block_num}: {iterations} stims with {rest_duration}s rest")
        
        for stim_in_block_count in range(iterations): 
            t0 = round(current_time * 100) / 100
            
            if t0 > max_time:
                print(f"Skipping remaining stims - T0 ({t0}s) beyond max time {max_time}s")
                return T0_PAIRS, BLOCK_INFO
            
            T0_PAIRS.append(t0)
            total_stim_count += 1
            BLOCK_INFO.append({
                'block': block_num,
                'stim_in_block': stim_in_block_count + 1, 
                'rest_duration': rest_duration,
                'total_stim_number': total_stim_count
            })
            
            current_time += stim_duration + rest_duration
    
    print(f"Calculated {len(T0_PAIRS)} total stimulations across {len(protocol_def['blocks'])} blocks.")
    return T0_PAIRS, BLOCK_INFO

def find_peaks_and_relaxation_times(df, t0_pairs, block_info_list): 
    """Detect peaks and calculate relaxation times for fatigue protocol"""
    df_processed = df.copy() 
    
    df_processed['Stim Peak'] = 0
    df_processed['True Peak Force (μN)'] = np.nan
    df_processed['BK corrected force (μN)'] = np.nan
    df_processed['Normalized Force (% of T1)'] = np.nan
    
    df_processed['Block Number'] = np.nan
    df_processed['Stim in Block'] = np.nan
    df_processed['Rest Duration (s)'] = np.nan
    df_processed['Stim Number'] = np.nan
    
    for metric_prefix in ['R10', 'R50', 'R80', 'R90']:
        df_processed[f'{metric_prefix} Reached'] = 0
        df_processed[f'{metric_prefix} Time'] = np.nan
        df_processed[f'{metric_prefix} Threshold Required (μN)'] = np.nan
    
    for metric_prefix in ['TT50P', 'TTP90']:
        df_processed[f'{metric_prefix} Reached'] = 0
        df_processed[f'{metric_prefix} Time'] = np.nan
        df_processed[f'{metric_prefix} Threshold Required (μN)'] = np.nan
    
    df_processed['Baseline Force (μN)'] = np.nan
    df_processed['Baseline Drift From Start (μN)'] = np.nan
    df_processed['Baseline Drift Rate (μN/s)'] = np.nan
    
    if 'Active Twitch Force (μN)_Smoothed' not in df_processed.columns:
        print("Error: 'Active Twitch Force (μN)_Smoothed' column not found. Please smooth data first.")
        return df_processed 

    for well in df_processed['Well'].unique():
        well_data = df_processed[df_processed['Well'] == well] 
        first_baseline = None
        first_peak_time = None
        first_bk_force = None
        
        for i, t0_val in enumerate(t0_pairs): 
            if i >= len(block_info_list): 
                print(f"Warning: Index {i} out of bounds for block_info_list (length {len(block_info_list)}) for well {well}. Skipping.")
                break
            block_data_item = block_info_list[i] 
            
            baseline_window_df = well_data[(well_data['Time (seconds)'] >= t0_val - BASELINE_WINDOW) & 
                                         (well_data['Time (seconds)'] < t0_val)]
            
            if baseline_window_df.empty: 
                continue
                
            baseline_force = baseline_window_df['Active Twitch Force (μN)_Smoothed'].mean()
            
            peak_search_df = well_data[(well_data['Time (seconds)'] >= t0_val) & 
                                      (well_data['Time (seconds)'] <= t0_val + PEAK_SEARCH_WINDOW)]
            
            if peak_search_df.empty: 
                continue
            
            stim_peak_original_index = peak_search_df['Active Twitch Force (μN)_Smoothed'].idxmax()
            stim_peak_time_val = well_data.loc[stim_peak_original_index, 'Time (seconds)'] 
            
            df_processed.loc[stim_peak_original_index, 'Stim Peak'] = 1
            df_processed.loc[stim_peak_original_index, 'Block Number'] = block_data_item['block']
            df_processed.loc[stim_peak_original_index, 'Stim in Block'] = block_data_item['stim_in_block']
            df_processed.loc[stim_peak_original_index, 'Rest Duration (s)'] = block_data_item['rest_duration']
            df_processed.loc[stim_peak_original_index, 'Stim Number'] = block_data_item['total_stim_number']
            
            true_peak_force_val = well_data.loc[stim_peak_original_index, 'Active Twitch Force (μN)_Smoothed'] 
            df_processed.loc[stim_peak_original_index, 'True Peak Force (μN)'] = true_peak_force_val
            
            df_processed.loc[stim_peak_original_index, 'Baseline Force (μN)'] = baseline_force
            
            if first_baseline is None:
                first_baseline = baseline_force
                first_peak_time = stim_peak_time_val
            
            drift_val = baseline_force - first_baseline 
            df_processed.loc[stim_peak_original_index, 'Baseline Drift From Start (μN)'] = drift_val
            
            if first_peak_time is not None and stim_peak_time_val > first_peak_time:
                time_diff_for_drift = stim_peak_time_val - first_peak_time 
                if time_diff_for_drift > 0:
                    drift_rate_val = drift_val / time_diff_for_drift 
                    df_processed.loc[stim_peak_original_index, 'Baseline Drift Rate (μN/s)'] = drift_rate_val
            
            bk_corrected_force_val = true_peak_force_val - baseline_force 
            df_processed.loc[stim_peak_original_index, 'BK corrected force (μN)'] = bk_corrected_force_val
            
            if first_bk_force is None:
                first_bk_force = bk_corrected_force_val
            
            if first_bk_force is not None and first_bk_force > 0:
                normalized_force_val = (bk_corrected_force_val / first_bk_force) * 100 
                df_processed.loc[stim_peak_original_index, 'Normalized Force (% of T1)'] = normalized_force_val
            
            relaxation_params = {
                'R10': (bk_corrected_force_val * 0.9) + baseline_force,
                'R50': (bk_corrected_force_val * 0.5) + baseline_force,
                'R80': (bk_corrected_force_val * 0.2) + baseline_force,
                'R90': (bk_corrected_force_val * 0.1) + baseline_force
            }
            
            relaxation_window_df = well_data[well_data['Time (seconds)'] > stim_peak_time_val] 
            
            for r_name, r_thresh_val in relaxation_params.items(): 
                df_processed.loc[stim_peak_original_index, f'{r_name} Threshold Required (μN)'] = r_thresh_val
                consecutive_count = 0
                r_start_idx = None
                
                for idx, row_data in relaxation_window_df.iterrows(): 
                    if row_data['Active Twitch Force (μN)_Smoothed'] <= (r_thresh_val * 1.02):
                        if consecutive_count == 0:
                            r_start_idx = idx
                        consecutive_count += 1
                        if consecutive_count == 3:
                            df_processed.loc[r_start_idx, f'{r_name} Reached'] = 1
                            time_diff_val = df_processed.loc[r_start_idx, 'Time (seconds)'] - stim_peak_time_val 
                            df_processed.loc[stim_peak_original_index, f'{r_name} Time'] = round(time_diff_val * 100) / 100
                            break
                    else:
                        consecutive_count = max(0, consecutive_count - 1) 
                
                if pd.isna(df_processed.loc[stim_peak_original_index, f'{r_name} Time']):
                    below_threshold_df = relaxation_window_df[
                        relaxation_window_df['Active Twitch Force (μN)_Smoothed'] <= r_thresh_val] 
                    
                    if not below_threshold_df.empty:
                        first_idx = below_threshold_df.index[0]
                        time_diff_val = below_threshold_df.loc[first_idx, 'Time (seconds)'] - stim_peak_time_val
                        df_processed.loc[stim_peak_original_index, f'{r_name} Time'] = round(time_diff_val * 100) / 100
                        df_processed.loc[stim_peak_original_index, f'{r_name} Reached'] = 1 
            
            tt50p_thresh_val = relaxation_params['R50'] 
            df_processed.loc[stim_peak_original_index, 'TT50P Threshold Required (μN)'] = tt50p_thresh_val
            
            contraction_window_df = well_data[(well_data['Time (seconds)'] >= t0_val) & 
                                             (well_data['Time (seconds)'] <= stim_peak_time_val)] 
            
            consecutive_count = 0
            tt50p_start_idx = None
            for idx, row_data in contraction_window_df.iterrows():
                if row_data['Active Twitch Force (μN)_Smoothed'] >= (tt50p_thresh_val * 0.98):
                    if consecutive_count == 0:
                        tt50p_start_idx = idx
                    consecutive_count += 1
                    if consecutive_count == 3:
                        df_processed.loc[tt50p_start_idx, 'TT50P Reached'] = 1
                        time_diff_val = df_processed.loc[tt50p_start_idx, 'Time (seconds)'] - t0_val
                        df_processed.loc[stim_peak_original_index, 'TT50P Time'] = round(time_diff_val * 100) / 100
                        break
                else:
                    consecutive_count = max(0, consecutive_count - 1) 
            
            if pd.isna(df_processed.loc[stim_peak_original_index, 'TT50P Time']):
                above_threshold_df = contraction_window_df[
                    contraction_window_df['Active Twitch Force (μN)_Smoothed'] >= tt50p_thresh_val] 
                if not above_threshold_df.empty:
                    first_idx = above_threshold_df.index[0]
                    time_diff_val = above_threshold_df.loc[first_idx, 'Time (seconds)'] - t0_val
                    df_processed.loc[stim_peak_original_index, 'TT50P Time'] = round(time_diff_val * 100) / 100
                    df_processed.loc[stim_peak_original_index, 'TT50P Reached'] = 1 

            ttp90_thresh_val = (bk_corrected_force_val * 0.9) + baseline_force 
            df_processed.loc[stim_peak_original_index, 'TTP90 Threshold Required (μN)'] = ttp90_thresh_val
            consecutive_count = 0
            ttp90_start_idx = None
            for idx, row_data in contraction_window_df.iterrows():
                if row_data['Active Twitch Force (μN)_Smoothed'] >= (ttp90_thresh_val * 0.98):
                    if consecutive_count == 0:
                        ttp90_start_idx = idx
                    consecutive_count += 1
                    if consecutive_count == 3:
                        df_processed.loc[ttp90_start_idx, 'TTP90 Reached'] = 1
                        time_diff_val = df_processed.loc[ttp90_start_idx, 'Time (seconds)'] - t0_val
                        df_processed.loc[stim_peak_original_index, 'TTP90 Time'] = round(time_diff_val * 100) / 100
                        break
                else:
                    consecutive_count = max(0, consecutive_count - 1) 
            
            if pd.isna(df_processed.loc[stim_peak_original_index, 'TTP90 Time']):
                above_threshold_df = contraction_window_df[
                    contraction_window_df['Active Twitch Force (μN)_Smoothed'] >= ttp90_thresh_val]
                if not above_threshold_df.empty:
                    first_idx = above_threshold_df.index[0]
                    time_diff_val = above_threshold_df.loc[first_idx, 'Time (seconds)'] - t0_val
                    df_processed.loc[stim_peak_original_index, 'TTP90 Time'] = round(time_diff_val * 100) / 100
                    df_processed.loc[stim_peak_original_index, 'TTP90 Reached'] = 1 
    return df_processed

def process_continuous_waveforms(df_input, plate_map_df, plate_name_str, max_time=300): 
    """Process raw continuous waveform data"""
    relevant_columns = ['Time (seconds)'] + [f'{r}{c} - Active Twitch Force (μN)' 
                                            for r in 'ABCD' for c in range(1, 7)]
    df_filtered = df_input[df_input.columns.intersection(relevant_columns)].copy() 
    
    if 'Time (seconds)' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['Time (seconds)'] <= max_time]
    
    for col_name in df_filtered.columns: 
        if col_name != 'Time (seconds)':
            df_filtered = df_filtered[df_filtered[col_name] != 0]

    if df_filtered.empty: 
        print(f"Warning: DataFrame became empty after filtering zero force values for {plate_name_str}")
        return pd.DataFrame()

    if 'Time (seconds)' in df_filtered.columns: 
        df_filtered['Time (seconds)'] = (df_filtered['Time (seconds)'] * 100).round() / 100
    
    data_cols_to_melt = [col for col in df_filtered.columns if col != 'Time (seconds)' and not df_filtered[col].isnull().all()] 
    
    if not data_cols_to_melt:
        print(f"Warning: No valid data columns found after filtering for {plate_name_str}")
        return pd.DataFrame()
        
    melted_df_output = df_filtered.melt(id_vars='Time (seconds)', value_vars=data_cols_to_melt, 
                                     var_name='Well_FullStr', value_name='Active Twitch Force (μN)') 
    
    melted_df_output['Well'] = melted_df_output['Well_FullStr'].str.split(' - ').str[0]
    
    def get_condition_val(well_code): 
        try:
            condition_str = plate_map_df.loc[well_code[0], well_code[1:]] 
            return condition_str
        except KeyError:
            return 'Unknown'
    
    melted_df_output['Condition'] = melted_df_output['Well'].apply(get_condition_val)
    
    if not melted_df_output.empty:
        all_common_time_points = None 
        for well_id_str in melted_df_output['Well'].unique(): 
            well_time_pts = set(melted_df_output[melted_df_output['Well'] == well_id_str]['Time (seconds)']) 
            if all_common_time_points is None:
                all_common_time_points = well_time_pts
            else:
                all_common_time_points = all_common_time_points.intersection(well_time_pts)
        
        if not all_common_time_points:
            print(f"Warning: No common time points found across wells for {plate_name_str}")
        
        if all_common_time_points: 
            melted_df_output = melted_df_output[melted_df_output['Time (seconds)'].isin(list(all_common_time_points))]

    if melted_df_output.empty: 
        print(f"Warning: DataFrame for {plate_name_str} is empty after time synchronization.")
        return pd.DataFrame()
        
    well_stats_df = melted_df_output.groupby('Well')['Active Twitch Force (μN)'].agg(['min', 'max']).reset_index() 
    melted_df_output = melted_df_output.merge(well_stats_df, on='Well', how='left') 
    
    melted_df_output['range_val'] = melted_df_output['max'] - melted_df_output['min'] 
    melted_df_output['Normalized Trace'] = np.where(
        melted_df_output['range_val'] > 1e-9, 
        (melted_df_output['Active Twitch Force (μN)'] - melted_df_output['min']) / melted_df_output['range_val'],
        0
    )
    
    melted_df_output = melted_df_output.drop(columns=['min', 'max', 'range_val', 'Well_FullStr']) 
    melted_df_output['Plate'] = plate_name_str
    
    melted_df_output = smooth_data(melted_df_output, 'Active Twitch Force (μN)', window_size=3)
    
    return melted_df_output

def create_fatigue_analysis_summary(data_list_input): 
    """Create comprehensive fatigue analysis focusing on block-by-block progression"""
    if not data_list_input:
        return pd.DataFrame(), pd.DataFrame()
        
    combined_data_df = pd.concat(data_list_input, ignore_index=True) 
    peaks_filtered_df = combined_data_df[combined_data_df['Stim Peak'] == 1].copy() 
    
    if peaks_filtered_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    
    detail_cols_list = ['Plate', 'Well', 'Condition', 'Time (seconds)', 'Stim Number', 
                       'Block Number', 'Stim in Block', 'Rest Duration (s)',
                       'True Peak Force (μN)', 'BK corrected force (μN)', 
                       'Normalized Force (% of T1)', 'Baseline Force (μN)',
                       'Baseline Drift From Start (μN)', 
                       'R10 Time', 'R50 Time', 'R80 Time', 'R90 Time', 
                       'TT50P Time', 'TTP90 Time'] 
    
    available_detail_cols_list = [col for col in detail_cols_list if col in peaks_filtered_df.columns] 
    fatigue_detail_output_df = peaks_filtered_df[available_detail_cols_list].copy() 
    
    if 'Time (seconds)' in fatigue_detail_output_df.columns:
        fatigue_detail_output_df = fatigue_detail_output_df.rename(columns={'Time (seconds)': 'Stim Peak Time (s)'}) 
    
    fatigue_detail_output_df = fatigue_detail_output_df.sort_values(['Plate', 'Well', 'Stim Number'])
    
    summary_data_output_list = [] 
    
    for (plate_val, well_val, condition_val), group_df in peaks_filtered_df.groupby(['Plate', 'Well', 'Condition']): 
        if group_df.empty: 
            continue
            
        group_df = group_df.sort_values('Stim Number')
        
        # --- CALCULATE FATIGUE RESISTANCE INDEX (FRI) ---
        # FRI is the mean of the Normalized Force across the ENTIRE protocol (Total)
        fri_total = np.nan
        if 'Normalized Force (% of T1)' in group_df.columns:
            # This takes the mean of all 40 stimulations (Total FRI)
            fri_total = group_df['Normalized Force (% of T1)'].mean()
            
        well_summary_item = { 
            'Plate': plate_val,
            'Well': well_val,
            'Condition': condition_val,
            # UPDATED NAME: Explicitly "Fatigue Resistance Index (%)"
            'Fatigue Resistance Index (%)': fri_total, 
            'Total Stims Detected': len(group_df) 
        }
        
        # Calculate FRI per block (Block Fatigue Resistance Index)
        if 'Block Number' in group_df.columns and 'Normalized Force (% of T1)' in group_df.columns:
             for block_num in sorted(group_df['Block Number'].unique()):
                 if pd.isna(block_num): continue
                 block_mask = group_df['Block Number'] == block_num
                 # This calculates the mean for just that block (e.g., Block 4 FRI)
                 block_fri = group_df.loc[block_mask, 'Normalized Force (% of T1)'].mean()
                 # UPDATED NAME: Explicitly "FRI Block X (%)"
                 well_summary_item[f'FRI Block {int(block_num)} (%)'] = block_fri

        # --- EXISTING METRICS ---
        if not group_df.empty:
            t1_data_row = group_df.iloc[0]
            if 'BK corrected force (μN)' in t1_data_row:
                well_summary_item['T1 BK Force (μN)'] = t1_data_row['BK corrected force (μN)']
            if 'Baseline Force (μN)' in t1_data_row:
                well_summary_item['T1 Baseline (μN)'] = t1_data_row['Baseline Force (μN)']
        
        if len(group_df) > 1: 
            last_data_row = group_df.iloc[-1]
            if 'BK corrected force (μN)' in last_data_row:
                well_summary_item['Last BK Force (μN)'] = last_data_row['BK corrected force (μN)']
            if 'Normalized Force (% of T1)' in last_data_row:
                well_summary_item['Last Normalized Force (% of T1)'] = last_data_row['Normalized Force (% of T1)']
            if 'Baseline Force (μN)' in last_data_row:
                well_summary_item['Last Baseline (μN)'] = last_data_row['Baseline Force (μN)']
                
                if 'T1 Baseline (μN)' in well_summary_item and pd.notna(well_summary_item.get('T1 Baseline (μN)')) and pd.notna(last_data_row['Baseline Force (μN)']):
                    drift_val = last_data_row['Baseline Force (μN)'] - well_summary_item['T1 Baseline (μN)']
                    well_summary_item['Total Baseline Drift (μN)'] = drift_val 
        
        # Add block rest info for context
        if 'Block Number' in group_df.columns:
            for block_num_val in sorted(group_df['Block Number'].unique()): 
                if pd.isna(block_num_val): continue
                block_data_df = group_df[group_df['Block Number'] == block_num_val] 
                if not block_data_df.empty:
                    rest_duration_val = block_data_df['Rest Duration (s)'].iloc[0] if 'Rest Duration (s)' in block_data_df.columns and not block_data_df['Rest Duration (s)'].empty else np.nan 
                    well_summary_item[f'Block {int(block_num_val)} Rest (s)'] = rest_duration_val
        
        if 'Normalized Force (% of T1)' in group_df.columns:
            norm_force_series = group_df['Normalized Force (% of T1)'].dropna()
            if len(norm_force_series) > 1 : 
                mean_val = norm_force_series.mean()
                if pd.notna(mean_val) and mean_val != 0: 
                    cv_val = (norm_force_series.std() / mean_val) * 100 
                    well_summary_item['Norm Force CV (%)'] = cv_val
            
        summary_data_output_list.append(well_summary_item)
    
    if not summary_data_output_list:
        return fatigue_detail_output_df, pd.DataFrame()
        
    fatigue_summary_output_df = pd.DataFrame(summary_data_output_list) 
    
    return fatigue_detail_output_df, fatigue_summary_output_df

def calculate_average_fatigue_progression(all_peak_data_df):
    """
    Calculates the average fatigue progression from all peak data, per condition.
    Averages 'Normalized Force (% of T1)' for each 'Condition' and 'Stim Number'.
    """
    if all_peak_data_df.empty:
        print("Warning: Peak data is empty, cannot calculate average fatigue progression.")
        return pd.DataFrame()
        
    required_cols = ['Condition', 'Stim Number', 'Normalized Force (% of T1)']
    if not all(col in all_peak_data_df.columns for col in required_cols):
        missing_cols = [col for col in required_cols if col not in all_peak_data_df.columns]
        print(f"Warning: Required columns {missing_cols} for average fatigue progression are missing.")
        return pd.DataFrame()

    peak_data_cleaned = all_peak_data_df.copy()
    peak_data_cleaned['Stim Number'] = pd.to_numeric(peak_data_cleaned['Stim Number'], errors='coerce')
    peak_data_cleaned['Normalized Force (% of T1)'] = pd.to_numeric(peak_data_cleaned['Normalized Force (% of T1)'], errors='coerce')
    # Condition is likely already string, but ensure it's not problematic for groupby
    peak_data_cleaned['Condition'] = peak_data_cleaned['Condition'].astype(str) 
    
    peak_data_cleaned = peak_data_cleaned.dropna(subset=['Condition', 'Stim Number', 'Normalized Force (% of T1)'])
    
    if peak_data_cleaned.empty: 
        print("Warning: No valid numeric data remaining after cleaning for average fatigue progression.")
        return pd.DataFrame()
    
    peak_data_cleaned['Stim Number'] = peak_data_cleaned['Stim Number'].astype(int)
    
    agg_functions = {
        'Normalized Force (% of T1)': ['mean', 'std', 'count']
    }
    # Group by Condition first, then Stim Number
    avg_progression = peak_data_cleaned.groupby(['Condition', 'Stim Number']).agg(agg_functions)
    
    avg_progression.columns = ['_'.join(col).strip() for col in avg_progression.columns.values]
    
    avg_progression = avg_progression.rename(columns={
        'Normalized Force (% of T1)_mean': 'Mean Normalized Force (% of T1)',
        'Normalized Force (% of T1)_std': 'StdDev Normalized Force (% of T1)',
        'Normalized Force (% of T1)_count': 'N (Wells contributing)'
    }).reset_index() # This will bring 'Condition' and 'Stim Number' back as columns
    
    return avg_progression

def calculate_condition_average_waveforms(list_of_processed_dfs): 
    """
    Calculates the average of 'Normalized Trace' for each 'Condition'
    from a list of processed waveform DataFrames.
    """
    if not list_of_processed_dfs:
        print("Warning: No processed data provided to calculate_condition_average_waveforms.")
        return pd.DataFrame()

    try:
        combined_waveforms = pd.concat(list_of_processed_dfs, ignore_index=True)
    except ValueError as e:
        print(f"Error concatenating waveform data lists for condition average: {e}.")
        return pd.DataFrame()

    if combined_waveforms.empty:
        print("Warning: Combined waveform data is empty for condition average.")
        return pd.DataFrame()

    required_cols = ['Time (seconds)', 'Normalized Trace', 'Condition']
    if not all(col in combined_waveforms.columns for col in required_cols):
        missing_cols = [col for col in required_cols if col not in combined_waveforms.columns]
        print(f"Warning: Required columns {missing_cols} for condition average waveform are missing.")
        return pd.DataFrame()

    combined_waveforms['Time (seconds)'] = (combined_waveforms['Time (seconds)'] * 100).round() / 100
    combined_waveforms['Normalized Trace'] = pd.to_numeric(combined_waveforms['Normalized Trace'], errors='coerce')
    
    combined_waveforms_cleaned = combined_waveforms.dropna(subset=['Time (seconds)', 'Normalized Trace', 'Condition'])

    if combined_waveforms_cleaned.empty:
        print("Warning: No valid data remaining after cleaning for condition average waveform.")
        return pd.DataFrame()

    agg_functions = {
        'Normalized Trace': ['mean', 'std', 'count']
    }
    avg_waveforms_by_condition = combined_waveforms_cleaned.groupby(['Condition', 'Time (seconds)']).agg(agg_functions)
    
    avg_waveforms_by_condition.columns = ['_'.join(col).strip() for col in avg_waveforms_by_condition.columns.values]
    
    avg_waveforms_by_condition = avg_waveforms_by_condition.rename(columns={
        'Normalized Trace_mean': 'Mean Normalized Trace (0-1)',
        'Normalized Trace_std': 'StdDev Normalized Trace (0-1)',
        'Normalized Trace_count': 'N (Wells contributing)'
    }).reset_index() 
    
    return avg_waveforms_by_condition


def generate_fatigue_reference(df_sample_input, t0_pairs_ref, block_info_ref, output_dir_path, day_str_ref, protocol_str_ref): 
    """Generate detailed reference document for fatigue protocol analysis"""
    
    example_well_str = None 
    example_data_row = None 
    
    if 'Stim Peak' in df_sample_input.columns: 
        peaks_df_ref = df_sample_input[df_sample_input['Stim Peak'] == 1] 
        if not peaks_df_ref.empty:
            first_peak_row_ref = peaks_df_ref.iloc[0] 
            example_well_str = first_peak_row_ref['Well']
            example_data_row = first_peak_row_ref
            print(f"Using well {example_well_str} for fatigue protocol reference")
    
    if example_data_row is None:
        print(f"No wells with peak data found for fatigue reference")
        return
    
    condition_str_ref = example_data_row['Condition'] 
    peak_time_val_ref = round(example_data_row['Time (seconds)'] * 100) / 100 
    
    reference_list = [ 
        f"# Fatigue Protocol Analysis - Reference Document ({day_str_ref}_{protocol_str_ref})",
        "",
        f"This document explains the fatigue protocol analysis methods with example calculations from Well {example_well_str}.",
        "",
        f"## Fatigue Protocol Structure",
        "",
        f"**Total Stimulations:** {len(t0_pairs_ref)}",
        f"**Protocol Blocks:**"
    ]
    
    for i, block_item_ref in enumerate(FATIGUE_PROTOCOL['blocks'], 1): 
        reference_list.append(f"- Block {i}: {block_item_ref['iterations']} stimulations with {block_item_ref['rest_duration']}s rest between")
    
    reference_list.extend([
        "",
        f"**Initial Delay:** {FATIGUE_PROTOCOL['initial_delay']}s",
        f"**Stimulation Duration:** {FATIGUE_PROTOCOL['stim_duration']}s (~75Hz for ~493ms)",
        "",
        f"## Example Calculation for Well {example_well_str} ({condition_str_ref})",
        "",
        f"### Basic Information",
        f"- Peak Time (detected): {peak_time_val_ref:.2f} seconds"
    ])
    
    if 'Block Number' in example_data_row:
        reference_list.append(f"- Block Number: {int(example_data_row['Block Number'])}")
    if 'Stim in Block' in example_data_row:
        reference_list.append(f"- Stimulation in Block: {int(example_data_row['Stim in Block'])}")
    if 'Rest Duration (s)' in example_data_row:
        reference_list.append(f"- Rest Duration: {example_data_row['Rest Duration (s)']}s")
    if 'Stim Number' in example_data_row:
        reference_list.append(f"- Overall Stimulation Number: {int(example_data_row['Stim Number'])}")
    
    reference_list.append("\n### Force Measurements")
    if 'True Peak Force (μN)' in example_data_row:
        reference_list.append(f"- True Peak Force: {example_data_row['True Peak Force (μN)']} μN")
    if 'Baseline Force (μN)' in example_data_row:
        reference_list.append(f"- Baseline Force: {example_data_row['Baseline Force (μN)']} μN")
    if 'BK corrected force (μN)' in example_data_row:
        reference_list.append(f"- BK Corrected Force: {example_data_row['BK corrected force (μN)']} μN")
    if 'Normalized Force (% of T1)' in example_data_row and pd.notna(example_data_row['Normalized Force (% of T1)']): # Check for NaN
        reference_list.append(f"- **Normalized Force (% of T1): {example_data_row['Normalized Force (% of T1)']:.2f}%** (KEY FATIGUE METRIC)")
    else:
        reference_list.append(f"- **Normalized Force (% of T1): N/A** (KEY FATIGUE METRIC)")


    reference_list.extend([
        "",
        "## Fatigue Analysis Methodology",
        "",
        "### 1. Protocol Timing Calculation",
        "- Parse fatigue protocol structure to calculate all T0 times",
        "- Account for variable rest durations between blocks",
        "- Track block information for each stimulation",
        "",
        "### 2. Waveform Processing (`process_continuous_waveforms`)",
        "- Filters raw data for relevant time and force columns.",
        "- Reshapes data from wide to long format.",
        "- Assigns experimental conditions based on the provided plate map.",
        "- Normalizes traces per well (0-1 scaling of raw force, creating `Normalized Trace`).",
        "- Applies a rolling window smoothing to `Active Twitch Force (μN)` creating `Active Twitch Force (μN)_Smoothed`.",
        "- Rows with zero force values in individual force columns are handled by attempting to keep non-zero numeric values.",
        "",
        "### 3. Peak Detection and Parameter Calculation (`find_peaks_and_relaxation_times`)",
        f"- For each T0, a baseline is calculated from smoothed force in a `{BASELINE_WINDOW}`s window immediately preceding T0.",
        f"- A peak is searched in smoothed force within a `{PEAK_SEARCH_WINDOW}`s window starting from T0.",
        "- **Baseline-Corrected (BK) Force:** `True Peak Force - Baseline Force` for that stimulation.",
        "- **Normalized Force (% of T1):** `(Current BK Force / T1 BK Force) * 100`. T1 is the first valid peak.",
        "- **Relaxation Times (e.g., R10, R50, R80, R90):** Time from peak for force to decay to X% of (Peak - Baseline) relative to baseline. Calculated from smoothed force.",
        "  - Detection requires 3 consecutive points below threshold (with 2% tolerance) or first crossing if 3 points not found.",
        "- **Contraction Times (e.g., TT50P, TTP90):** Time from T0 for force to reach X% of (Peak - Baseline) relative to baseline, during the rising phase. Calculated from smoothed force.",
        "  - Detection requires 3 consecutive points above threshold (with 2% tolerance) or first crossing if 3 points not found.",
        "- **Baseline Drift:** Tracks changes in baseline force over the protocol duration.",
        "",
        "### 4. Summary Generation (`create_fatigue_analysis_summary`)",
        "- **Fatigue Resistance Index (%)**: The Mean Normalized Force (%) across the entire protocol. Higher = Better resistance.",
        "- **Fatigue Detail Sheet:** Contains per-stimulation data for all detected peaks, including all calculated time-based metrics (R10-R90 Time, TT50P-TTP90 Time).",
        "- **Fatigue Summary Sheet:** Aggregates metrics per well, including FRI, T1 values, last stimulation values, block-wise averages of normalized force.",
        "- **Condition Avg Fatigue Sheet:** Mean, StdDev, and N for 'Normalized Force (% of T1)' vs. 'Stim Number', grouped by 'Condition'.",
        "- **Condition Avg Waveforms Sheet:** Mean, StdDev, and N for 'Normalized Trace (0-1 per well)' vs. 'Time (seconds)', grouped by 'Condition'.",
        ""
    ])
    
    current_time_ref = FATIGUE_PROTOCOL['initial_delay'] 
    for i, block_item_ref in enumerate(FATIGUE_PROTOCOL['blocks'], 1):
        reference_list.append(f"**Block {i} (Rest: {block_item_ref['rest_duration']}s)**")
        reference_list.append(f"- Stimulations: {(i-1)*10 + 1} to {i*10}") 
        block_duration_ref = block_item_ref['iterations'] * (FATIGUE_PROTOCOL['stim_duration'] + block_item_ref['rest_duration']) - block_item_ref['rest_duration'] 
        reference_list.append(f"- Time range: ~{current_time_ref:.1f}s to ~{current_time_ref + block_duration_ref:.1f}s")
        reference_list.append("")
        current_time_ref += block_item_ref['iterations'] * (FATIGUE_PROTOCOL['stim_duration'] + block_item_ref['rest_duration'])
    
    reference_list.extend([
        "## Expected Fatigue Patterns",
        "",
        "### Normal Healthy Response:",
        "- Block 1 (1.5s rest): Slight decline ~85-100% of T1",
        "- Block 2 (1.0s rest): Moderate decline ~60-80% of T1", 
        "- Block 3 (0.5s rest): Significant decline of T1",
        "- Block 4 (0.25s rest): Significant decline of T1",
        "",
        "### Fatigue-Resistant Response:",
        "- Minimal decline across all blocks",
        "- May show potentiation (>100%) in later blocks",
        "",
        "### Fatigue-Sensitive Response:",
        "- Early and pronounced decline",
        "- May reach <30% of T1 in final blocks",
        "",
        "## Data Interpretation Guidelines",
        "",
        "### Primary Endpoints:",
        "1. **Fatigue Resistance Index (FRI)**: Primary metric for overall performance.",
        "2. **Final Force (% of T1)**: Overall fatigue resistance",
        "3. **Block 4 Mean (% of T1)**: Performance under minimal rest",
        "4. **Condition Avg Fatigue / Condition Avg Waveforms**: Visual representation of fatigue trends per condition.",
        "",
        "### Secondary Endpoints:",
        "1. **Force CV (%)**: Consistency across stimulations", 
        "2. **Baseline Drift**: Distinguish from true force fatigue",
        "3. **Relaxation/Contraction Times (from Fatigue Detail sheet)**: Individual time metrics like R50, R90, TT50P, TTP90.",
        "",
        "## Statistical Considerations",
        "",
        "- Compare FRI and block means between conditions",
        "- Consider both magnitude and timing of fatigue onset",
        "- Account for baseline drift in interpretation",
        "",
        "## Analysis Parameters",
        "",
        f"- Baseline window: {BASELINE_WINDOW}s before each T0",
        f"- Peak search window: {PEAK_SEARCH_WINDOW}s after each T0", 
        f"- Detection algorithm: 3 consecutive points with 2% tolerance",
        f"- Time precision: 0.01s (100Hz sampling rate)",
        "",
        "## Note on Protocol Flexibility",
        "",
        "This analysis automatically adapts to the fatigue protocol structure.",
        "T0 times are calculated from the protocol definition, ensuring accurate",
        "peak detection even with variable rest durations between blocks."
    ])
    
    reference_file_path = os.path.join(output_dir_path, f"{day_str_ref}_{protocol_str_ref}_fatigue_reference.md") 
    try:
        with open(reference_file_path, 'w', encoding='utf-8') as f_out: 
            f_out.write('\n'.join(reference_list))
        print(f"Fatigue protocol reference written to {reference_file_path}")
    except Exception as e_write: 
        print(f"Error writing fatigue reference: {e_write}")

def process_and_combine(input_files, plate_map, output_dir, max_recording_time):
    processed_data = {}
    average_traces = {}

    for input_file in tqdm(input_files, desc="Processing files"):
        try:
            # --- CSV/Excel Detection & Loading ---
            if input_file.lower().endswith('.csv'):
                try:
                    raw_df = pd.read_csv(input_file)
                    if 'Time (s)' in raw_df.columns:
                        raw_df.rename(columns={'Time (s)': 'Time (seconds)'}, inplace=True)
                    
                    continuous_waveforms_df = pd.DataFrame()
                    if 'Time (seconds)' in raw_df.columns:
                        continuous_waveforms_df['Time (seconds)'] = raw_df['Time (seconds)']
                    
                    for col in raw_df.columns:
                        if col in ['Time (seconds)', 'Unnamed: 0']: continue
                        new_col_name = f"{col} - Active Twitch Force (μN)"
                        continuous_waveforms_df[new_col_name] = raw_df[col]
                except Exception as e:
                    print(f"Error processing CSV {input_file}: {e}")
                    continue
            else:
                xls = pd.ExcelFile(input_file)
                if 'continuous-waveforms' not in xls.sheet_names:
                    # print(f"'continuous-waveforms' sheet not found in {input_file}. Skipping.")
                    continue
                continuous_waveforms_df = pd.read_excel(xls, 'continuous-waveforms')
            
            # Round time values
            if 'Time (seconds)' in continuous_waveforms_df.columns:
                continuous_waveforms_df['Time (seconds)'] = (continuous_waveforms_df['Time (seconds)'] * 100).round() / 100
                
        except Exception as e:
            print(f"Error reading input file {input_file}: {e}")
            continue

        filename = os.path.basename(input_file)
        file_info = parse_filename(filename)
        if not file_info:
            # print(f"Could not parse filename: {filename}")
            continue
            
        day, protocol, plate_name = file_info
        
        # [CRITICAL FIX] Strip extension
        if protocol and '.' in protocol: 
            protocol = protocol.split('.')[0]

        continuous_df = process_continuous_waveforms(continuous_waveforms_df, plate_map, plate_name, max_recording_time)
        
        if continuous_df.empty:
            # print(f"No data processed for {filename}. Skipping.")
            continue
        
        # Find peaks using Fatigue logic
        continuous_df = find_peaks_and_relaxation_times(continuous_df, T0_PAIRS, BLOCK_INFO)

        if (day, protocol) not in processed_data:
            processed_data[(day, protocol)] = {}
        
        for condition in continuous_df['Condition'].unique():
            if condition not in processed_data[(day, protocol)]:
                processed_data[(day, protocol)][condition] = []
            condition_df = continuous_df[continuous_df['Condition'] == condition]
            processed_data[(day, protocol)][condition].append(condition_df)

    # --- Output Generation ---
    for (day, protocol), conditions in processed_data.items():
        average_traces[(day, protocol)] = {}
        for condition, data_list in conditions.items():
            # Use Condition Average Waveforms logic
            average_traces[(day, protocol)][condition] = calculate_condition_average_waveforms(data_list)

        # Output file
        output_filename = f"{day}_{protocol}_FatigueAnalysis.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        print(f"Creating workbook for {day}_{protocol}")
        wb = Workbook()
        wb.remove(wb.active)  
        
        all_peak_data = [] # Collect all peaks for summary
        
        # 1. Condition Sheets (Full Data)
        for condition, data_list in conditions.items():
            if not data_list: continue
            combined_continuous = pd.concat(data_list, ignore_index=True)
            
            # Add Condition Sheet
            df_to_excel_sheet(wb, f"{condition}", combined_continuous)
            
            all_peak_data.extend(data_list) # Collect for summary

        # 2. Create Fatigue Summaries
        if all_peak_data:
            fatigue_detail_df, fatigue_summary_df = create_fatigue_analysis_summary(all_peak_data)
            
            if not fatigue_summary_df.empty:
                df_to_excel_sheet(wb, "Fatigue Summary", fatigue_summary_df)
            
            if not fatigue_detail_df.empty:
                df_to_excel_sheet(wb, "Fatigue Detail", fatigue_detail_df)
                
            # 3. Condition Avg Fatigue (New Sheet)
            combined_all_data = pd.concat(all_peak_data, ignore_index=True)
            peaks_only = combined_all_data[combined_all_data['Stim Peak'] == 1]
            avg_fatigue_progression = calculate_average_fatigue_progression(peaks_only)
            if not avg_fatigue_progression.empty:
                df_to_excel_sheet(wb, "Condition Avg Fatigue", avg_fatigue_progression)

        # 4. Condition Avg Waveforms (Representative Traces)
        avg_waves_list = []
        for cond in conditions.keys():
            if cond in average_traces[(day, protocol)]:
                avg_wave = average_traces[(day, protocol)][cond]
                if not avg_wave.empty:
                    avg_wave['Condition'] = cond
                    avg_waves_list.append(avg_wave)
        
        if avg_waves_list:
            all_avg_waves = pd.concat(avg_waves_list, ignore_index=True)
            df_to_excel_sheet(wb, "Condition Avg Waveforms", all_avg_waves)

        print(f"Saving workbook to {output_path}")
        wb.save(output_path)
        
        # Generate Reference
        if all_peak_data:
             # Just grab the first dataframe to extract sample data
             sample_df = all_peak_data[0]
             generate_fatigue_reference(sample_df, T0_PAIRS, BLOCK_INFO, output_dir, day, protocol)

def main():
    """Main function for fatigue protocol analysis"""
    
    print("\n*** Fatigue Protocol Analysis v1.1.3 ***") 
    print("Analyzing 4-block fatigue protocol with decreasing rest periods")
    print("Protocol: 3s delay + 4 blocks (10 stims each) with 1.5s→1.0s→0.5s→0.25s rest")
    print(f"Configuration: Max recording time set to {MAX_RECORDING_TIME}s")
    
    # Use global MAX_RECORDING_TIME
    global T0_PAIRS, BLOCK_INFO 
    T0_PAIRS, BLOCK_INFO = calculate_fatigue_protocol_times(FATIGUE_PROTOCOL, MAX_RECORDING_TIME)
    
    if not T0_PAIRS:
        print("Error: Could not calculate fatigue protocol timing. Exiting.")
        return
    
    print(f"\nCalculated {len(T0_PAIRS)} stimulations across {len(FATIGUE_PROTOCOL['blocks'])} blocks")
    
    input_files = select_files("Select Fatigue input Excel or CSV files", [
        ("Data files", "*.xlsx *.csv"),  # <--- This is now the default (shows both)
        ("Excel files", "*.xlsx"), 
        ("CSV files", "*.csv"), 
        ("All files", "*.*")
    ])
    if not input_files:
        print("No input files selected. Exiting.")
        return
    
    plate_map_file_selected = select_files("Select plate map CSV file", [("CSV files", "*.csv")]) 
    if not plate_map_file_selected:
        print("No plate map file selected. Exiting.")
        return
    plate_map_path_str = plate_map_file_selected[0] 

    output_dir_selected = select_directory("Select output directory for fatigue analysis results") 
    if not output_dir_selected:
        print("No output directory selected. Exiting.")
        return

    try:
        plate_map_input_df = pd.read_csv(plate_map_path_str, index_col=0, encoding='utf-8') 
        plate_map_input_df.index = plate_map_input_df.index.astype(str)
        plate_map_input_df.columns = plate_map_input_df.columns.astype(str)
    except Exception as e_load_map: 
        print(f"Error loading plate map: {e_load_map}")
        return

    process_and_combine(input_files, plate_map_input_df, output_dir_selected, MAX_RECORDING_TIME)
    
    print("\nFatigue protocol analysis complete!")
    print("\nKey outputs:")
    print("- Fatigue Detail: Individual stimulation data with block information and R/TT times.")
    print("- Fatigue Summary: Block-by-block analysis and overall fatigue metrics.")
    print("- Condition Avg Fatigue: Mean and StdDev of Normalized Force (% of T1) vs. Stim Number, per condition.")
    print("- Condition Avg Waveforms: Mean and StdDev of Normalized Trace (0-1 per well) vs. Time (s), per condition.")
    print("- Condition sheets: Complete waveform data for visualization")
    print("- Reference document: Detailed methodology and interpretation guide")

if __name__ == "__main__":
    main()