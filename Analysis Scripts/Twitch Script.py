# -*- coding: utf-8 -*-
"""
Created on Mon May 19 2025

@author: pb92
@version: 1.2.2

CHANGELOG:
v1.2.2 - INTERFACE UPDATE: Moved all interactive parameters (num_stims, thresholds) to the
         global CONFIGURATION SECTION. Removed console prompts for "Click and Run" simplicity.
       - ADDED: CSV support for Mantarray Controller files.
v1.2.1 - Enhanced Visualizations & Artifact Rejection (Fast Peak/Min Threshold).
"""

import os
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from tqdm import tqdm
import uuid

# ==========================================
#        CONFIGURATION SECTION
# ==========================================

# Recording Settings
MAX_RECORDING_TIME = 65.0        # Maximum time (seconds) to process.
NUM_STIMS = 10                   # Number of stimulations to analyze.

# Twitch Protocol Settings (0.2Hz Default)
INITIAL_DELAY = 4.97             # Time to first stimulation (seconds)
STIM_DURATION = 0.01             # Pulse duration (seconds)
REST_DURATION = 4.99             # Time between pulses (seconds)

# Analysis Parameters
BASELINE_WINDOW = 0.2            # Window before T0 for baseline calculation
PEAK_SEARCH_WINDOW = 0.7         # Window after T0 to search for peak

# Artifact Rejection
MINIMUM_TWITCH_THRESHOLD_UN = 5.0  # Min BK corrected force (µN) for valid twitch
FAST_PEAK_WARNING_S = 0.04         # Time to Peak (from T0) threshold for warning

# Visualization Settings
VIS_SUBPLOT_FIXED_HEIGHT_INCHES = 3.5
VIS_MAX_PEAKS_PER_PDF_PAGE = 5
VIS_WINDOW_BEFORE_T0 = 0.2
VIS_WINDOW_AFTER_T0 = 0.8

# ==========================================
#      END CONFIGURATION SECTION
# ==========================================

T0_PAIRS = []  # Will store stimulation start times

def select_files(title, file_types):
    root = tk.Tk()
    root.withdraw()
    files = filedialog.askopenfilenames(title=title, filetypes=file_types)
    root.destroy()
    return files

def select_directory(title):
    root = tk.Tk()
    root.withdraw()
    directory = filedialog.askdirectory(title=title)
    root.destroy()
    return directory

def df_to_excel_sheet(wb, sheet_name, df, start_cell='A1'):
    df_formatted = df.copy()
    for col in df_formatted.columns:
        if 'Time' in col and col != 'Time (seconds)' and df_formatted[col].dtype in ['float64', 'float32']:
            if any(kinetic_term in col for kinetic_term in ['R10 Time', 'R50 Time', 'R80 Time', 'R90 Time', 'TT50P Time', 'TTP90 Time']):
                 df_formatted[col] = df_formatted[col].round(3)
            else:
                 df_formatted[col] = df_formatted[col].round(2)
        elif 'Force' in col or 'Threshold' in col:
             df_formatted[col] = df_formatted[col].round(2)
        elif col == 'Relative Time (ms)':
            df_formatted[col] = df_formatted[col].round(0)
    
    ws = wb.create_sheet(title=sheet_name)
    rows = dataframe_to_rows(df_formatted, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    if df_formatted.empty: return

    max_row = len(df_formatted) + 1
    max_col = len(df_formatted.columns)
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    safe_table_name = f"Table_{uuid.uuid4().hex[:8]}"
    
    try:
        tab = Table(displayName=safe_table_name, ref=f"{start_cell}:{end_cell}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except ValueError as e:
        print(f"Warning: Could not create table in sheet '{sheet_name}': {str(e)}")
    
    max_sample_rows = 100
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column[:min(max_sample_rows, len(df_formatted)+1)]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max(10, min(max_length + 2, 50))
        ws.column_dimensions[column_letter].width = adjusted_width

def parse_filename(filename):
    parts = os.path.basename(filename).split('_')
    day = next((part for part in parts if part.lower().startswith('d') and part[1:].isdigit()), None)
    if day:
        try:
            day_index = parts.index(day)
            if day_index + 1 < len(parts) and day_index + 2 < len(parts):
                plate_name = parts[day_index + 1]
                protocol_part = parts[day_index + 2]
                protocol = protocol_part.split('.')[0] if '.' in protocol_part else protocol_part
                return day, protocol, plate_name
        except ValueError:
            pass
    print(f"Warning: Could not parse day, protocol, and plate name from filename: {filename}")
    return None, None, None

def smooth_data(df, column, window_size=3):
    df[f'{column}_Smoothed'] = df.groupby('Well')[column].rolling(
        window=window_size, center=True, min_periods=1).mean().reset_index(0, drop=True)
    return df

def calculate_stimulation_times(initial_delay, stim_duration, rest_duration, num_stims, max_time):
    global T0_PAIRS
    T0_PAIRS = []
    for i in range(num_stims):
        t0 = initial_delay + i * (stim_duration + rest_duration)
        t0 = round(t0 * 100) / 100
        if t0 > max_time:
            print(f"Skipping stim {i+1} at time {t0} (beyond max time {max_time})")
            break
        T0_PAIRS.append(t0)
    print(f"Calculated {len(T0_PAIRS)} stimulation times.")
    return T0_PAIRS

def find_peaks_and_relaxation_times(df, t0_pairs, min_twitch_threshold, fast_peak_warning_s):
    df = df.copy()
    # Initialize columns
    cols_to_add = [
        'Stim Peak', 'True Peak Force (μN)', 'BK corrected force (μN)', 'Normalized Force (% of T1)', 'Fast Peak Warning',
        'R10 Reached', 'R10 Time', 'R10 Threshold Required (μN)',
        'R50 Reached', 'R50 Time', 'R50 Threshold Required (μN)',
        'R80 Reached', 'R80 Time', 'R80 Threshold Required (μN)',
        'R90 Reached', 'R90 Time', 'R90 Threshold Required (μN)',
        'TT50P Reached', 'TT50P Time', 'TT50P Threshold Required (μN)',
        'TTP90 Reached', 'TTP90 Time', 'TTP90 Threshold Required (μN)',
        'Baseline Force (μN)', 'Baseline Drift From Start (μN)', 'Baseline Drift Rate (μN/s)', 'Stim Number'
    ]
    for col in cols_to_add:
        df[col] = np.nan if 'Reached' not in col and 'Warning' not in col and 'Peak' not in col else 0

    for well in df['Well'].unique():
        well_data = df[df['Well'] == well].copy()
        first_baseline = None
        first_peak_time_abs = None
        first_valid_bk_force = None
        stim_count = 0
        
        for i, t0 in enumerate(t0_pairs):
            baseline_window_df = well_data[(well_data['Time (seconds)'] >= t0 - BASELINE_WINDOW) & 
                                           (well_data['Time (seconds)'] < t0)]
            if baseline_window_df.empty: continue
                
            baseline = baseline_window_df['Active Twitch Force (μN)_Smoothed'].mean()
            
            peak_search_df = well_data[(well_data['Time (seconds)'] >= t0) & 
                                       (well_data['Time (seconds)'] <= t0 + PEAK_SEARCH_WINDOW)]
            if peak_search_df.empty: continue
            
            stim_peak_index = peak_search_df['Active Twitch Force (μN)_Smoothed'].idxmax()
            stim_peak_time_abs = well_data.loc[stim_peak_index, 'Time (seconds)']
            
            df.loc[stim_peak_index, 'Stim Peak'] = 1
            stim_count += 1
            df.loc[stim_peak_index, 'Stim Number'] = stim_count
            
            true_peak_force = well_data.loc[stim_peak_index, 'Active Twitch Force (μN)_Smoothed']
            df.loc[stim_peak_index, 'True Peak Force (μN)'] = true_peak_force
            df.loc[stim_peak_index, 'Baseline Force (μN)'] = baseline
            
            if first_baseline is None:
                first_baseline = baseline
                first_peak_time_abs = stim_peak_time_abs
            
            drift = baseline - first_baseline
            df.loc[stim_peak_index, 'Baseline Drift From Start (μN)'] = drift
            
            if first_peak_time_abs is not None and stim_peak_time_abs > first_peak_time_abs:
                time_diff_for_drift = stim_peak_time_abs - first_peak_time_abs
                if time_diff_for_drift > 0:
                    drift_rate = drift / time_diff_for_drift
                    df.loc[stim_peak_index, 'Baseline Drift Rate (μN/s)'] = drift_rate
            
            bk_corrected_force = true_peak_force - baseline
            
            time_to_peak_from_t0 = stim_peak_time_abs - t0
            if time_to_peak_from_t0 < fast_peak_warning_s:
                df.loc[stim_peak_index, 'Fast Peak Warning'] = 1

            if bk_corrected_force < min_twitch_threshold:
                df.loc[stim_peak_index, 'BK corrected force (μN)'] = np.nan
                if first_valid_bk_force is None and stim_count == 1:
                    first_valid_bk_force = np.nan
                if pd.notna(first_valid_bk_force) and first_valid_bk_force > 0:
                    normalized_force = (bk_corrected_force / first_valid_bk_force) * 100
                    df.loc[stim_peak_index, 'Normalized Force (% of T1)'] = normalized_force
                elif first_valid_bk_force is None and stim_count > 1:
                     pass
                else:
                    df.loc[stim_peak_index, 'Normalized Force (% of T1)'] = np.nan
                continue

            df.loc[stim_peak_index, 'BK corrected force (μN)'] = bk_corrected_force

            if first_valid_bk_force is None:
                first_valid_bk_force = bk_corrected_force
            
            if pd.notna(first_valid_bk_force) and first_valid_bk_force > 0:
                normalized_force = (bk_corrected_force / first_valid_bk_force) * 100
                df.loc[stim_peak_index, 'Normalized Force (% of T1)'] = normalized_force
            else:
                df.loc[stim_peak_index, 'Normalized Force (% of T1)'] = np.nan

            r10_threshold = (bk_corrected_force * 0.9) + baseline
            r50_threshold = (bk_corrected_force * 0.5) + baseline
            r80_threshold = (bk_corrected_force * 0.2) + baseline
            r90_threshold = (bk_corrected_force * 0.1) + baseline
            
            df.loc[stim_peak_index, 'R10 Threshold Required (μN)'] = r10_threshold
            df.loc[stim_peak_index, 'R50 Threshold Required (μN)'] = r50_threshold
            df.loc[stim_peak_index, 'R80 Threshold Required (μN)'] = r80_threshold
            df.loc[stim_peak_index, 'R90 Threshold Required (μN)'] = r90_threshold
            
            relaxation_window_df = well_data[well_data['Time (seconds)'] > stim_peak_time_abs].copy() 
            
            for r_name, r_threshold in [('R10', r10_threshold), ('R50', r50_threshold), 
                                        ('R80', r80_threshold), ('R90', r90_threshold)]:
                consecutive_count = 0
                r_start_idx = None
                for idx_loop, row_loop in relaxation_window_df.iterrows():
                    if row_loop['Active Twitch Force (μN)_Smoothed'] <= (r_threshold * 1.02):
                        if consecutive_count == 0: r_start_idx = idx_loop
                        consecutive_count += 1
                        if consecutive_count == 3:
                            time_diff_relax = well_data.loc[r_start_idx, 'Time (seconds)'] - stim_peak_time_abs
                            time_diff_relax = round(time_diff_relax * 1000) / 1000
                            df.loc[stim_peak_index, f'{r_name} Time'] = time_diff_relax
                            df.loc[stim_peak_index, f'{r_name} Reached'] = 1
                            break 
                    else:
                        consecutive_count = max(0, consecutive_count - 1)
                
                if pd.isna(df.loc[stim_peak_index, f'{r_name} Time']):
                    below_threshold_df = relaxation_window_df[relaxation_window_df['Active Twitch Force (μN)_Smoothed'] <= r_threshold]
                    if not below_threshold_df.empty:
                        first_idx_fallback = below_threshold_df.index[0]
                        time_diff_relax = below_threshold_df.loc[first_idx_fallback, 'Time (seconds)'] - stim_peak_time_abs
                        time_diff_relax = round(time_diff_relax * 1000) / 1000
                        df.loc[stim_peak_index, f'{r_name} Time'] = time_diff_relax
                        df.loc[stim_peak_index, f'{r_name} Reached'] = 1
            
            tt50p_threshold = r50_threshold 
            df.loc[stim_peak_index, 'TT50P Threshold Required (μN)'] = tt50p_threshold
            contraction_window_df = well_data[(well_data['Time (seconds)'] >= t0) & 
                                              (well_data['Time (seconds)'] <= stim_peak_time_abs)].copy()
            
            consecutive_count = 0
            tt50p_start_idx = None
            for idx_loop, row_loop in contraction_window_df.iterrows():
                if row_loop['Active Twitch Force (μN)_Smoothed'] >= (tt50p_threshold * 0.98):
                    if consecutive_count == 0: tt50p_start_idx = idx_loop
                    consecutive_count += 1
                    if consecutive_count == 3:
                        time_diff_contract = well_data.loc[tt50p_start_idx, 'Time (seconds)'] - t0
                        time_diff_contract = round(time_diff_contract * 1000) / 1000
                        df.loc[stim_peak_index, 'TT50P Time'] = time_diff_contract
                        df.loc[stim_peak_index, 'TT50P Reached'] = 1
                        break
                else:
                    consecutive_count = max(0, consecutive_count - 1)
            
            if pd.isna(df.loc[stim_peak_index, 'TT50P Time']):
                above_threshold_df = contraction_window_df[contraction_window_df['Active Twitch Force (μN)_Smoothed'] >= tt50p_threshold]
                if not above_threshold_df.empty:
                    first_idx_fallback = above_threshold_df.index[0]
                    time_diff_contract = above_threshold_df.loc[first_idx_fallback, 'Time (seconds)'] - t0
                    time_diff_contract = round(time_diff_contract * 1000) / 1000
                    df.loc[stim_peak_index, 'TT50P Time'] = time_diff_contract
                    df.loc[stim_peak_index, 'TT50P Reached'] = 1

            ttp90_threshold = (bk_corrected_force * 0.9) + baseline 
            df.loc[stim_peak_index, 'TTP90 Threshold Required (μN)'] = ttp90_threshold
            consecutive_count = 0
            ttp90_start_idx = None
            for idx_loop, row_loop in contraction_window_df.iterrows():
                if row_loop['Active Twitch Force (μN)_Smoothed'] >= (ttp90_threshold * 0.98):
                    if consecutive_count == 0: ttp90_start_idx = idx_loop
                    consecutive_count += 1
                    if consecutive_count == 3:
                        time_diff_contract = well_data.loc[ttp90_start_idx, 'Time (seconds)'] - t0
                        time_diff_contract = round(time_diff_contract * 1000) / 1000
                        df.loc[stim_peak_index, 'TTP90 Time'] = time_diff_contract
                        df.loc[stim_peak_index, 'TTP90 Reached'] = 1
                        break
                else:
                    consecutive_count = max(0, consecutive_count - 1)
            
            if pd.isna(df.loc[stim_peak_index, 'TTP90 Time']):
                above_threshold_df = contraction_window_df[contraction_window_df['Active Twitch Force (μN)_Smoothed'] >= ttp90_threshold]
                if not above_threshold_df.empty:
                    first_idx_fallback = above_threshold_df.index[0]
                    time_diff_contract = above_threshold_df.loc[first_idx_fallback, 'Time (seconds)'] - t0
                    time_diff_contract = round(time_diff_contract * 1000) / 1000
                    df.loc[stim_peak_index, 'TTP90 Time'] = time_diff_contract
                    df.loc[stim_peak_index, 'TTP90 Reached'] = 1
    return df

def extract_normalized_twitches(df):
    twitches = []
    for (plate, well, condition) in df[['Plate', 'Well', 'Condition']].drop_duplicates().values:
        subset = df[(df['Plate'] == plate) & (df['Well'] == well) & (df['Condition'] == condition)]
        stim_peak_times = subset[subset['Stim Peak'] == 1]['Time (seconds)']
        
        for peak_time in stim_peak_times:
            peak_time = round(peak_time * 100) / 100
            start_time = peak_time - 1
            end_time = peak_time + 1
            twitch_data = subset[(subset['Time (seconds)'] >= start_time) & 
                                (subset['Time (seconds)'] <= end_time)].copy()
            
            if not twitch_data.empty:
                twitch_data['Relative Time (ms)'] = ((twitch_data['Time (seconds)'] - peak_time) * 1000).round(0)
                min_force = twitch_data['Active Twitch Force (μN)_Smoothed'].min()
                max_force = twitch_data['Active Twitch Force (μN)_Smoothed'].max()
                
                if max_force > min_force:
                    twitch_data['Normalized Force (%)'] = ((twitch_data['Active Twitch Force (μN)_Smoothed'] - min_force) / 
                                                          (max_force - min_force)) * 100
                else:
                    twitch_data['Normalized Force (%)'] = 0
                
                twitches.append(twitch_data)
    
    if not twitches: return pd.DataFrame()
    result = pd.concat(twitches, ignore_index=True)
    return result.sort_values(['Plate', 'Well', 'Condition', 'Time (seconds)'])

def process_continuous_waveforms(df, plate_map, plate_name, day_val, max_time):
    relevant_columns = ['Time (seconds)'] + [f'{row}{col} - Active Twitch Force (μN)' 
                                            for row in 'ABCD' for col in range(1, 7)]
    df_filtered = df[df.columns.intersection(relevant_columns)].copy()
    
    if 'Time (seconds)' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['Time (seconds)'] <= max_time]
    
    if 'Time (seconds)' in df_filtered.columns:
        df_filtered.loc[:, 'Time (seconds)'] = (df_filtered['Time (seconds)'] * 100).round() / 100
    
    data_columns = [col for col in df_filtered.columns if col != 'Time (seconds)' and not df_filtered[col].isnull().all()]
    
    if not data_columns: return pd.DataFrame()
        
    melted_df = df_filtered.melt(id_vars='Time (seconds)', value_vars=data_columns, 
                                var_name='Well_Raw', value_name='Active Twitch Force (μN)')
    
    melted_df['Well'] = melted_df['Well_Raw'].str.split(' - ').str[0]
    melted_df.drop(columns=['Well_Raw'], inplace=True)
    
    def get_condition(well_code):
        try:
            return plate_map.loc[well_code[0], str(well_code[1:])]
        except KeyError:
            return 'Unknown'
    
    melted_df['Condition'] = melted_df['Well'].apply(get_condition)
    
    well_stats = melted_df.groupby('Well')['Active Twitch Force (μN)'].agg(['min', 'max']).rename(columns={'min':'well_min', 'max':'well_max'})
    melted_df = melted_df.merge(well_stats, left_on='Well', right_index=True)
    
    melted_df['range'] = melted_df['well_max'] - melted_df['well_min']
    melted_df['Normalized Trace'] = np.where(
        melted_df['range'] > 0,
        (melted_df['Active Twitch Force (μN)'] - melted_df['well_min']) / melted_df['range'],
        0
    )
    
    melted_df = melted_df.drop(columns=['well_min', 'well_max', 'range'])
    melted_df['Plate'] = plate_name
    melted_df['Day'] = day_val
    
    melted_df = smooth_data(melted_df, 'Active Twitch Force (μN)', window_size=3)
    
    return melted_df

def calculate_average_trace(data_list):
    if not data_list:
        return pd.DataFrame(columns=['Time (seconds)', 'Normalized Trace', 'Standard Deviation'])
        
    combined_data = pd.concat(data_list, ignore_index=True)
    combined_data['Time (seconds)'] = (combined_data['Time (seconds)'] * 100).round() / 100
    
    grouped = combined_data.groupby('Time (seconds)')
    avg_trace = grouped['Normalized Trace'].mean().reset_index()
    std_trace = grouped['Normalized Trace'].std().reset_index()
    avg_trace['Standard Deviation'] = std_trace['Normalized Trace']
    
    if 'Active Twitch Force (μN)_Smoothed' in combined_data.columns:
        avg_smoothed_force = grouped['Active Twitch Force (μN)_Smoothed'].mean().reset_index(name='Avg Smoothed Force (μN)')
        std_smoothed_force = grouped['Active Twitch Force (μN)_Smoothed'].std().reset_index(name='Smoothed Force StdDev (μN)')
        avg_trace = pd.merge(avg_trace, avg_smoothed_force, on='Time (seconds)', how='left')
        avg_trace = pd.merge(avg_trace, std_smoothed_force, on='Time (seconds)', how='left')

    return avg_trace

def create_baseline_drift_summary(data_list):
    if not data_list: return pd.DataFrame(), pd.DataFrame()
    combined_data = pd.concat(data_list, ignore_index=True)
    peaks_df = combined_data[combined_data['Stim Peak'] == 1].copy()
    if peaks_df.empty: return pd.DataFrame(), pd.DataFrame()
    
    required_columns = ['Day', 'Plate', 'Well', 'Condition', 'Time (seconds)', 'Stim Number',
                        'Baseline Force (μN)', 'Baseline Drift From Start (μN)', 
                        'Baseline Drift Rate (μN/s)', 'BK corrected force (μN)']
    available_columns = [col for col in required_columns if col in peaks_df.columns]
    drift_detail = peaks_df[available_columns].copy()
    
    if 'Baseline Drift Rate (μN/s)' in drift_detail.columns:
        drift_detail.rename(columns={'Baseline Drift Rate (μN/s)': 'Baseline Drift Rate from start (μN/s)'}, inplace=True)
    if 'Time (seconds)' in drift_detail.columns:
        drift_detail.rename(columns={'Time (seconds)': 'Stim Peak Time'}, inplace=True)
    
    sort_columns = ['Day', 'Plate', 'Well', 'Stim Number'] if 'Stim Number' in drift_detail.columns else ['Day', 'Plate', 'Well']
    drift_detail.sort_values(sort_columns, inplace=True)
    
    # Normalized drift logic
    first_twitch_data = {}
    if all(col in drift_detail.columns for col in ['Stim Number', 'Baseline Drift From Start (μN)', 'BK corrected force (μN)']):
        for group_keys, group in drift_detail.groupby(['Day', 'Plate', 'Well', 'Condition']):
            valid_stims = group.dropna(subset=['BK corrected force (μN)'])
            if not valid_stims.empty:
                t1_data = valid_stims[valid_stims['Stim Number'] == valid_stims['Stim Number'].min()]
                if not t1_data.empty:
                    first_twitch_data[group_keys] = {'bk_corrected_force': t1_data['BK corrected force (μN)'].iloc[0]}
        
        normalized_drifts = []
        for idx, row in drift_detail.iterrows():
            key = (row['Day'], row['Plate'], row['Well'], row['Condition'])
            if key in first_twitch_data and pd.notna(row.get('Baseline Drift From Start (μN)')):
                first_twitch_force = first_twitch_data[key]['bk_corrected_force']
                if pd.notna(first_twitch_force) and first_twitch_force > 0:
                    normalized_drifts.append((row['Baseline Drift From Start (μN)'] / first_twitch_force) * 100)
                else: normalized_drifts.append(np.nan)
            else: normalized_drifts.append(np.nan)
        drift_detail['Normalized Drift (% of Valid T1 Force)'] = normalized_drifts
    
    well_summary = []
    for group_keys_summary, group in drift_detail.groupby(['Day', 'Plate', 'Well', 'Condition']):
        if 'Stim Number' in group.columns: group.sort_values('Stim Number', inplace=True)
        summary = {'Day': group_keys_summary[0], 'Plate': group_keys_summary[1], 'Well': group_keys_summary[2], 'Condition': group_keys_summary[3]}
        
        key_for_t1 = (summary['Day'], summary['Plate'], summary['Well'], summary['Condition'])
        if key_for_t1 in first_twitch_data:
            summary['Valid T1 BK Corrected Force (μN)'] = first_twitch_data[key_for_t1]['bk_corrected_force']

        if 'Stim Number' in group.columns and 'Baseline Force (μN)' in group.columns and not group.empty:
            t1_data = group[group['Stim Number'] == group['Stim Number'].min()]
            if not t1_data.empty:
                t1_baseline = t1_data['Baseline Force (μN)'].iloc[0]
                summary['T1 Baseline (μN)'] = t1_baseline
                last_data = group.iloc[-1]
                last_baseline = last_data['Baseline Force (μN)']
                last_stim_num = int(last_data['Stim Number']) if pd.notna(last_data['Stim Number']) else 'N/A'
                summary[f'Last T Baseline (T{last_stim_num}) (μN)'] = last_baseline
                if pd.notna(t1_baseline):
                    last_drift = last_baseline - t1_baseline
                    summary[f'Drift Last-T1 (μN)'] = last_drift
                    if summary.get('Valid T1 BK Corrected Force (μN)', 0) > 0:
                        summary['Normalized Drift Last (% of Valid T1 Force)'] = (last_drift / summary['Valid T1 BK Corrected Force (μN)']) * 100
                if pd.notna(t1_baseline) and t1_baseline > 0:
                    summary['Baseline Ratio Last/T1'] = last_baseline / t1_baseline
                if 'Stim Peak Time' in last_data and 'Stim Peak Time' in t1_data.iloc[0]:
                    time_diff = last_data['Stim Peak Time'] - t1_data['Stim Peak Time'].iloc[0]
                    if pd.notna(time_diff) and time_diff > 0:
                        summary['Drift Rate (μN/s)'] = last_drift / time_diff
        
        if 'Baseline Force (μN)' in group.columns and len(group) > 1:
            min_b, max_b = group['Baseline Force (μN)'].min(), group['Baseline Force (μN)'].max()
            summary['Baseline Range (μN)'] = max_b - min_b
            if summary.get('Valid T1 BK Corrected Force (μN)', 0) > 0:
                summary['Normalized Baseline Range (% of Valid T1 Force)'] = ((max_b - min_b) / summary['Valid T1 BK Corrected Force (μN)']) * 100
        
        if 'Baseline Drift From Start (μN)' in group.columns and 'Stim Number' in group.columns:
            if not group['Baseline Drift From Start (μN)'].empty and group['Baseline Drift From Start (μN)'].notna().any():
                max_drift_row = group.loc[group['Baseline Drift From Start (μN)'].idxmax()]
                max_drift = max_drift_row['Baseline Drift From Start (μN)']
                summary['Max Drift (μN)'] = max_drift
                summary['Max Drift Stim #'] = int(max_drift_row['Stim Number']) if pd.notna(max_drift_row['Stim Number']) else 'N/A'
                if summary.get('Valid T1 BK Corrected Force (μN)', 0) > 0:
                    summary['Normalized Max Drift (% of Valid T1 Force)'] = (max_drift / summary['Valid T1 BK Corrected Force (μN)']) * 100
        well_summary.append(summary)
    
    if not well_summary: return drift_detail, pd.DataFrame()
    drift_summary = pd.DataFrame(well_summary)
    # Column ordering logic...
    return drift_detail, drift_summary

def create_contractile_summary(data_list):
    if not data_list: return pd.DataFrame()
    combined_data = pd.concat(data_list, ignore_index=True)
    peaks_df = combined_data[(combined_data['Stim Peak'] == 1) & (combined_data['BK corrected force (μN)'].notna())].copy()
    if peaks_df.empty: return pd.DataFrame()
    
    agg_dict = {}
    if 'BK corrected force (μN)' in peaks_df.columns: agg_dict['BK corrected force (μN)'] = ['mean', 'max', 'std']
    for time_col in ['R10 Time', 'R50 Time', 'R80 Time', 'R90 Time', 'TT50P Time', 'TTP90 Time']:
        if time_col in peaks_df.columns: agg_dict[time_col] = 'mean'
    if 'Stim Number' in peaks_df.columns: agg_dict['Stim Number'] = 'count'
    
    if not agg_dict: return pd.DataFrame()
    
    grouping_cols = ['Day', 'Plate', 'Well', 'Condition']
    force_summary = peaks_df.groupby(grouping_cols).agg(agg_dict).reset_index()
    force_summary.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in force_summary.columns]
    
    rename_dict = {
        'BK corrected force (μN) mean': 'Mean BK Corrected Force (μN)',
        'BK corrected force (μN) max': 'Max BK Corrected Force (μN)',
        'BK corrected force (μN) std': 'BK Corrected Force StdDev (μN)',
        'R10 Time mean': 'R10 Time Avg (s)', 'R50 Time mean': 'R50 Time Avg (s)',
        'R80 Time mean': 'R80 Time Avg (s)', 'R90 Time mean': 'R90 Time Avg (s)',
        'TT50P Time mean': 'TT50P Time Avg (s)', 'TTP90 Time mean': 'TTP90 Time Avg (s)',
        'Stim Number count': 'Valid Twitch Count'
    }
    force_summary.rename(columns={k: v for k, v in rename_dict.items() if k in force_summary.columns}, inplace=True)
    
    if 'Mean BK Corrected Force (μN)' in force_summary.columns:
        force_summary['BK Force CV (%)'] = np.where(force_summary['Mean BK Corrected Force (μN)'] != 0,
            (force_summary['BK Corrected Force StdDev (μN)'] / force_summary['Mean BK Corrected Force (μN)']) * 100, np.nan)

    # Fatigue Index Logic
    early, late, fatigue = [], [], []
    for idx, row in force_summary.iterrows():
        group_data = peaks_df[(peaks_df['Day'] == row['Day']) & (peaks_df['Plate'] == row['Plate']) & 
                              (peaks_df['Well'] == row['Well']) & (peaks_df['Condition'] == row['Condition'])]
        if 'Stim Number' in group_data.columns:
            sorted_group = group_data.sort_values('Stim Number')
            e, l, f = np.nan, np.nan, np.nan
            if len(sorted_group) >= 5: e = sorted_group.head(5)['BK corrected force (μN)'].mean()
            if len(sorted_group) >= 10: l = sorted_group.tail(5)['BK corrected force (μN)'].mean()
            if pd.notna(e) and pd.notna(l) and e > 0: f = l / e
            early.append(e); late.append(l); fatigue.append(f)
        else:
            early.append(np.nan); late.append(np.nan); fatigue.append(np.nan)
            
    force_summary['Early Twitch Mean (T1-T5) (μN)'] = early
    force_summary['Late Twitch Mean (Last 5) (μN)'] = late
    force_summary['Fatigue Index (Late/Early)'] = fatigue
    
    return force_summary

def create_relaxation_summary(data_list):
    if not data_list: return pd.DataFrame()
    combined_data = pd.concat(data_list, ignore_index=True)
    summary = combined_data[combined_data['Stim Peak'] == 1].copy()
    if summary.empty: return pd.DataFrame()
    
    available_columns = ['Day', 'Plate', 'Well', 'Condition', 'Time (seconds)', 'Stim Number', 'Fast Peak Warning']
    kinetic_cols = ['True Peak Force (μN)', 'BK corrected force (μN)', 'Baseline Force (μN)', 'Normalized Force (% of T1)', 
                    'Baseline Drift From Start (μN)', 'Baseline Drift Rate (μN/s)',
                    'R10 Time', 'R10 Threshold Required (μN)', 'R10 Reached',
                    'R50 Time', 'R50 Threshold Required (μN)', 'R50 Reached',
                    'R80 Time', 'R80 Threshold Required (μN)', 'R80 Reached',
                    'R90 Time', 'R90 Threshold Required (μN)', 'R90 Reached',
                    'TT50P Time', 'TT50P Threshold Required (μN)', 'TT50P Reached',
                    'TTP90 Time', 'TTP90 Threshold Required (μN)', 'TTP90 Reached']
    available_columns.extend([col for col in kinetic_cols if col in summary.columns])
    summary_detail = summary[available_columns].copy()
    if 'Time (seconds)' in summary_detail.columns: summary_detail.rename(columns={'Time (seconds)': 'Stim Peak Time'}, inplace=True)
    sort_cols = ['Day', 'Plate', 'Well', 'Stim Number'] if 'Stim Number' in summary_detail.columns else ['Day', 'Plate', 'Well']
    return summary_detail.sort_values(sort_cols)

def visualize_peak_detection(df_vis, output_dir_path, day_id, protocol_id, t0_list_global, peak_search_window_global):
    try:
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
    except ImportError:
        print("Matplotlib not available. Skipping visualization.")
        return

    required_cols_vis = ['Day', 'Plate', 'Well', 'Condition', 'Time (seconds)', 
                         'Active Twitch Force (μN)', 'Active Twitch Force (μN)_Smoothed', 'Stim Peak', 'Stim Number']
    if not all(col in df_vis.columns for col in required_cols_vis): return

    pdf_output_path = os.path.join(output_dir_path, f"{day_id}_{protocol_id}_peak_detection.pdf")
    with PdfPages(pdf_output_path) as pdf_file:
        for (plate_name_vis, well_code, condition_vis), well_plate_data_group in df_vis.groupby(['Plate', 'Well', 'Condition']):
            peaks_in_well = well_plate_data_group[well_plate_data_group['Stim Peak'] == 1].copy()
            if peaks_in_well.empty: continue 
            
            peaks_in_well.sort_values('Stim Number', inplace=True)
            total_peaks = len(peaks_in_well)
            num_pages = (total_peaks + VIS_MAX_PEAKS_PER_PDF_PAGE - 1) // VIS_MAX_PEAKS_PER_PDF_PAGE

            for page_num in range(num_pages):
                start_idx = page_num * VIS_MAX_PEAKS_PER_PDF_PAGE
                end_idx = min(start_idx + VIS_MAX_PEAKS_PER_PDF_PAGE, total_peaks)
                peaks_page = peaks_in_well.iloc[start_idx:end_idx]
                if len(peaks_page) == 0: continue

                fig, axes = plt.subplots(len(peaks_page), 1, figsize=(12, len(peaks_page) * VIS_SUBPLOT_FIXED_HEIGHT_INCHES), squeeze=False)
                fig.suptitle(f"Peak Detection: {day_id} {protocol_id} - {plate_name_vis} - {well_code} ({condition_vis})", fontsize=12, y=0.99)

                for i_subplot, (peak_row_idx, peak_row) in enumerate(peaks_page.iterrows()):
                    ax = axes[i_subplot, 0] 
                    peak_time = peak_row['Time (seconds)']
                    stim_num = int(peak_row['Stim Number']) - 1
                    t0_val = t0_list_global[stim_num] if 0 <= stim_num < len(t0_list_global) else peak_time - 0.3
                    
                    plot_start = t0_val - VIS_WINDOW_BEFORE_T0
                    plot_end = t0_val + VIS_WINDOW_AFTER_T0
                    trace = well_plate_data_group[(well_plate_data_group['Time (seconds)'] >= plot_start) & (well_plate_data_group['Time (seconds)'] <= plot_end)]
                    
                    if 'Active Twitch Force (μN)' in trace.columns: ax.plot(trace['Time (seconds)'], trace['Active Twitch Force (μN)'], 'b-', alpha=0.4, label='Raw')
                    if 'Active Twitch Force (μN)_Smoothed' in trace.columns: ax.plot(trace['Time (seconds)'], trace['Active Twitch Force (μN)_Smoothed'], 'k-', label='Smoothed')
                    
                    ax.axvline(x=t0_val, color='lime', linestyle='--', label='T0')
                    ax.axvspan(t0_val, t0_val + peak_search_window_global, alpha=0.1, color='gold', label='Search Win')
                    ax.axvline(x=peak_time, color='r', linestyle='--', label='Peak')
                    
                    if pd.notna(peak_row.get('Baseline Force (μN)')): ax.axhline(y=peak_row['Baseline Force (μN)'], color='gray', linestyle=':', label='Baseline')
                    
                    if pd.notna(peak_row.get('R50 Threshold Required (μN)')):
                        ax.axhline(y=peak_row['R50 Threshold Required (μN)'], color='orange', linestyle='-.', label='R50 Th')
                        if peak_row.get('R50 Reached') == 1: ax.plot(peak_time + peak_row['R50 Time'], peak_row['R50 Threshold Required (μN)'], 'o', color='orange')

                    bk_val = peak_row.get('BK corrected force (μN)')
                    txt = f"BK Force: {bk_val:.2f}µN" if pd.notna(bk_val) else "BK Force: < Threshold"
                    color = 'lightgreen' if pd.notna(bk_val) else 'lightcoral'
                    ax.text(0.02, 0.95, txt, transform=ax.transAxes, bbox=dict(boxstyle='round', fc=color, alpha=0.6))
                    
                    ax.set_title(f"Stim #{int(peak_row['Stim Number'])}")
                    ax.legend(loc='upper right', fontsize='x-small')
                
                plt.tight_layout(rect=[0, 0.03, 1, 0.96])
                pdf_file.savefig(fig)
                plt.close(fig)
    print(f"Visualizations saved to {pdf_output_path}")

def generate_calculation_reference(df, t0_pairs_ref, output_dir, day, protocol, min_twitch_thresh_ref, fast_peak_warn_s_ref, peak_search_win_ref):
    example_peak_row = None
    if 'Stim Peak' in df.columns:
        valid_peaks = df[(df['Stim Peak'] == 1) & (df['BK corrected force (μN)'].notna())]
        if not valid_peaks.empty: example_peak_row = valid_peaks.sort_values('Stim Number').iloc[0]
        else:
            any_peaks = df[df['Stim Peak'] == 1]
            if not any_peaks.empty: example_peak_row = any_peaks.sort_values('Stim Number').iloc[0]

    md_content = [
        f"# 1Hz Analysis Reference ({day}_{protocol})",
        f"- Min Twitch Threshold: {min_twitch_thresh_ref:.2f} uN",
        f"- Fast Peak Warning: {fast_peak_warn_s_ref:.3f} s",
        f"- Initial Delay: {INITIAL_DELAY} s",
        f"- Stim Duration: {STIM_DURATION} s",
        f"- Rest Duration: {REST_DURATION} s"
    ]
    
    if example_peak_row is not None:
        md_content.append(f"\n## Example: Well {example_peak_row['Well']}")
        md_content.append(f"- Peak Force: {example_peak_row.get('True Peak Force (μN)', 0):.2f} uN")
        md_content.append(f"- Baseline: {example_peak_row.get('Baseline Force (μN)', 0):.2f} uN")
        md_content.append(f"- BK Force: {example_peak_row.get('BK corrected force (μN)', 0):.2f} uN")
    
    ref_path = os.path.join(output_dir, f"{day}_{protocol}_reference_v1.2.2.md")
    try:
        with open(ref_path, 'w', encoding='utf-8') as f: f.write('\n'.join(md_content))
        print(f"Reference doc saved: {ref_path}")
    except Exception as e: print(f"Error saving reference: {e}")

def process_and_combine(input_files, plate_map, output_dir, num_stims, max_recording_time, min_twitch_thresh, fast_peak_s, peak_search_win):
    processed_data = {}
    
    for input_file in tqdm(input_files, desc="Processing files"):
        try:
            if input_file.lower().endswith('.csv'):
                raw_df = pd.read_csv(input_file)
                if 'Time (s)' in raw_df.columns: raw_df.rename(columns={'Time (s)': 'Time (seconds)'}, inplace=True)
                continuous_waveforms_df_raw = pd.DataFrame()
                if 'Time (seconds)' in raw_df.columns: continuous_waveforms_df_raw['Time (seconds)'] = raw_df['Time (seconds)']
                for col in raw_df.columns:
                    if col not in ['Time (seconds)', 'Unnamed: 0']: continuous_waveforms_df_raw[f"{col} - Active Twitch Force (μN)"] = raw_df[col]
            else:
                xls = pd.ExcelFile(input_file)
                if 'continuous-waveforms' not in xls.sheet_names: continue
                continuous_waveforms_df_raw = pd.read_excel(xls, 'continuous-waveforms')
            
            if 'Time (seconds)' in continuous_waveforms_df_raw.columns:
                continuous_waveforms_df_raw['Time (seconds)'] = (continuous_waveforms_df_raw['Time (seconds)'] * 100).round() / 100
                
        except Exception as e:
            print(f"Error reading {input_file}: {e}")
            continue

        day_val, protocol_val, plate_name_val = parse_filename(os.path.basename(input_file))
        if protocol_val and '.' in protocol_val: protocol_val = protocol_val.split('.')[0]
        
        if not (day_val and protocol_val and plate_name_val): continue
            
        continuous_df_processed = process_continuous_waveforms(continuous_waveforms_df_raw, plate_map, plate_name_val, day_val, max_recording_time)
        if continuous_df_processed.empty: continue
        
        continuous_df_kinetics = find_peaks_and_relaxation_times(continuous_df_processed, T0_PAIRS, min_twitch_thresh, fast_peak_s)
        
        key = (day_val, protocol_val)
        if key not in processed_data: processed_data[key] = []
        processed_data[key].append(continuous_df_kinetics)

    for (day, proto), dfs in processed_data.items():
        if not dfs: continue
        combined = pd.concat(dfs, ignore_index=True)
        out_name = f"{day}_{proto}_kinetics_v1.2.2.xlsx"
        out_path = os.path.join(output_dir, out_name)
        
        print(f"\nCreating workbook for {day}_{proto}")
        wb = Workbook(); wb.remove(wb.active)
        
        if not combined.empty: df_to_excel_sheet(wb, "Full Processed Data", combined)
        
        drift_det, drift_sum = create_baseline_drift_summary([combined])
        if not drift_det.empty: df_to_excel_sheet(wb, "Baseline Drift Detail", drift_det)
        if not drift_sum.empty: df_to_excel_sheet(wb, "Baseline Drift Summary", drift_sum)
        
        relax_det = create_relaxation_summary([combined])
        if not relax_det.empty: df_to_excel_sheet(wb, "Relaxation Detail", relax_det)
        
        contract_avg = create_contractile_summary([combined])
        if not contract_avg.empty: df_to_excel_sheet(wb, "Contractile Averages", contract_avg)
        
        traces = []
        if 'Condition' in combined.columns:
            for cond, group in combined.groupby('Condition'):
                tr = calculate_average_trace([group])
                if not tr.empty:
                    tr['Condition'] = cond
                    traces.append(tr)
        if traces: df_to_excel_sheet(wb, "Representative Traces", pd.concat(traces, ignore_index=True))
        
        wb.save(out_path)
        print(f"Saved to {out_path}")
        
        if not combined.empty:
            visualize_peak_detection(combined, output_dir, day, proto, T0_PAIRS, peak_search_win)
            generate_calculation_reference(combined, T0_PAIRS, output_dir, day, proto, min_twitch_thresh, fast_peak_s, peak_search_win)

def main():
    print("\n*** 1Hz Protocol Twitch Analysis v1.2.2 ***")
    print(f"Configuration: {NUM_STIMS} stims, Max Time {MAX_RECORDING_TIME}s")
    
    calculate_stimulation_times(INITIAL_DELAY, STIM_DURATION, REST_DURATION, NUM_STIMS, MAX_RECORDING_TIME)
    
    input_files = select_files("Select input Excel or CSV files", [
        ("Data files", "*.xlsx *.csv"),
        ("Excel files", "*.xlsx"), 
        ("CSV files", "*.csv"), 
        ("All files", "*.*")
    ])
    if not input_files: return

    plate_map_path = select_files("Select plate map CSV file", [("CSV files", "*.csv")])
    if not plate_map_path: return
    plate_map_path = plate_map_path[0]

    output_dir = select_directory("Select output directory")
    if not output_dir: return

    try:
        plate_map = pd.read_csv(plate_map_path, index_col=0, encoding='utf-8')
        plate_map.index = plate_map.index.astype(str)
        plate_map.columns = plate_map.columns.astype(str)
    except Exception as e:
        print(f"Error loading plate map: {e}")
        return

    process_and_combine(input_files, plate_map, output_dir, NUM_STIMS, MAX_RECORDING_TIME,
                        MINIMUM_TWITCH_THRESHOLD_UN, FAST_PEAK_WARNING_S, PEAK_SEARCH_WINDOW)

if __name__ == "__main__":
    main()