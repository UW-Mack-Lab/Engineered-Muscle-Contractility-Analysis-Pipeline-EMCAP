# -*- coding: utf-8 -*-
"""
Created on Wed Jul 16 12:40:14 2025

@author: pb92
@version: 1.1.0 - Added CSV Support (Fixed)

CHANGELOG:
v1.1.0 - Added support for time force CSV files.
       - Script now detects .csv vs .xlsx and formats CSV data fomrats
       - Fixed file dialog to explicitly show *.csv files.
v1.0.5 - Finalized for Collaboration
       - MAX_RECORDING_TIME moved to global configuration parameters.
"""

import os
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog, simpledialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from tqdm import tqdm
import uuid

# ==========================================
#        CONFIGURATION SECTION
# ==========================================

# Recording Settings
MAX_RECORDING_TIME = 350.0    # Maximum time (seconds) to process. Trims data beyond this point.

# FvF Protocol Settings (Hard-coded from JSON)
INITIAL_DELAY = 21.0          # Initial delay before first stimulation
STIM_DURATION = 2.0           # Duration of each stimulation
INTER_STIM_DELAY = 16.0       # Delay between stimulations
BASELINE_WINDOW = 1.0         # Window before T0 to calculate baseline
FVF_PROTOCOL_HZ = [1, 2, 3, 5, 10, 15, 20, 30, 40, 60, 80, 100]  # Hz sequence

# ==========================================
#      END CONFIGURATION SECTION
# ==========================================

# Global storage for protocol triplets
T0_T1_HZ_TRIPLETS = []

def select_files(title, file_types):
    root = tk.Tk()
    root.withdraw()
    files = filedialog.askopenfilenames(title=title, filetypes=file_types)
    return files

def select_directory(title):
    root = tk.Tk()
    root.withdraw()
    directory = filedialog.askdirectory(title=title)
    return directory

def df_to_excel_sheet(wb, sheet_name, df, start_cell='A1'):
    # Format numbers with consistent decimal places before sending to Excel
    df_formatted = df.copy()
    
    # Apply formatting to numeric columns based on their content
    for col in df_formatted.columns:
        if 'Time' in col and col != 'Time (seconds)' and df_formatted[col].dtype in ['float64', 'float32']:
            # Time-related columns (like relaxation times) - 2 decimal places
            df_formatted[col] = df_formatted[col].round(2)
        elif 'Force' in col or 'Threshold' in col:
            # Force-related columns - maintain higher precision
            pass  # Don't round force values
        elif col == 'Relative Time (ms)':
            # Round to nearest millisecond
            df_formatted[col] = df_formatted[col].round(0)
        # Don't round Normalized Trace or Normalized Force (%) values
    
    ws = wb.create_sheet(title=sheet_name)
    
    rows = dataframe_to_rows(df_formatted, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    max_row = len(df_formatted) + 1  # +1 for header
    max_col = len(df_formatted.columns)
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    
    # Generate a random, unique table name to avoid the expensive check
    safe_table_name = f"Table_{uuid.uuid4().hex[:8]}"
    
    try:
        tab = Table(displayName=safe_table_name, ref=f"{start_cell}:{end_cell}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except ValueError as e:
        print(f"Warning: Could not create table in sheet '{sheet_name}': {str(e)}")
    
    # Optimize column width setting - only do a sample of rows
    max_sample_rows = 100
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        # Only check header and a sample of rows (faster)
        for cell in column[:min(max_sample_rows, len(df_formatted)+1)]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max(10, min(max_length + 2, 50))  # Between 10 and 50
        ws.column_dimensions[column_letter].width = adjusted_width

def parse_filename(filename):
    parts = filename.split('_')
    day = next((part for part in parts if part.lower().startswith('d') and part[1:].isdigit()), None)
    if day:
        try:
            day_index = parts.index(day)
            if day_index + 1 < len(parts) and day_index + 2 < len(parts):
                plate_name = parts[day_index + 1]
                protocol_part = parts[day_index + 2]
                # Strip extension if present (handles .xlsx and .csv)
                protocol = protocol_part.split('.')[0]
                return day, protocol, plate_name
        except IndexError:
            pass
    print(f"Warning: Could not parse day, protocol, and plate name from filename: {filename}")
    return None, None, None

def smooth_data(df, column, window_size=3):
    try:
        df[f'{column}_Smoothed'] = df.groupby('Well')[column].rolling(window=window_size, center=True, min_periods=1).mean().reset_index(0, drop=True)
        return df
    except Exception as e:
        print(f"Error in smooth_data function: {e}")
        # Create empty smoothed column if operation fails
        df[f'{column}_Smoothed'] = df[column]
        return df

def parse_fvf_protocol(max_time):
    """Parse Force vs Frequency (FvF) protocol using global constants"""
    global T0_T1_HZ_TRIPLETS
    
    print("\n=== Force vs Frequency (FvF) Protocol ===")
    print("Based on configuration parameters:")
    print(f"- Initial Delay: {INITIAL_DELAY}s")
    print(f"- Each Stimulation: {STIM_DURATION}s duration") 
    print(f"- Inter-stimulation Delay: {INTER_STIM_DELAY}s")
    print(f"- Baseline Window: {BASELINE_WINDOW}s before each T0")
    print("\nStimulation Schedule:")
    
    T0_T1_HZ_TRIPLETS = []
    current_time = INITIAL_DELAY
    
    for i, hz in enumerate(FVF_PROTOCOL_HZ):
        t0 = current_time
        t1 = t0 + STIM_DURATION
        
        # Round to 0.01s precision
        t0 = round(t0 * 100) / 100
        t1 = round(t1 * 100) / 100
        
        # Skip if beyond max recording time
        if t1 > max_time:
            print(f"Skipping stim {i+1} ({hz}Hz) at time {t1} (beyond max time {max_time})")
            break
            
        T0_T1_HZ_TRIPLETS.append((t0, t1, hz))
        print(f"  Stim {i+1:2d}: {t0:6.1f}s - {t1:6.1f}s ({hz:3d}Hz)")
        
        # Move to next stimulation (add stim duration + inter-stim delay)
        current_time = t1 + INTER_STIM_DELAY

    print(f"\nTotal stimulations to detect: {len(T0_T1_HZ_TRIPLETS)}")
    print("=" * 45)
    return T0_T1_HZ_TRIPLETS

def find_peaks_and_relaxation_times(df, t0_t1_hz_triplets):
    df = df.copy()
    df['Stim Peak'] = 0
    df['Stim Number'] = 0         # Stimulation number (1, 2, 3, ...)
    df['Hz'] = 0                  # Frequency for this stimulation (separate filterable column)
    df['Kinetics Peak Force (μN)'] = np.nan    # ±100ms around T1 for kinetics calculations
    df['Max Tetanic Force (μN)'] = np.nan      # T0 to T1+200ms search for maximum contractile capacity
    df['Baseline (μN)'] = np.nan
    df['Kinetics Peak-BK (μN)'] = np.nan       # Kinetics Peak - Baseline (for thresholds)
    df['Max tetanic-BK (μN)'] = np.nan         # Max Tetanic - Baseline
    # Add columns to track kinetics peak details
    df['Kinetics Peak Time (seconds)'] = np.nan
    df['Peak Time Delta (s)'] = np.nan
    df['R10 Reached'] = 0
    df['R10 Time'] = np.nan
    df['R10 Threshold Required (μN)'] = np.nan
    df['R50 Reached'] = 0
    df['R50 Time'] = np.nan
    df['R50 Threshold Required (μN)'] = np.nan
    df['R80 Reached'] = 0
    df['R80 Time'] = np.nan
    df['R80 Threshold Required (μN)'] = np.nan
    df['R90 Reached'] = 0
    df['R90 Time'] = np.nan
    df['R90 Threshold Required (μN)'] = np.nan
    df['TT50P Reached'] = 0
    df['TT50P Time'] = np.nan
    df['TT50P Threshold Required (μN)'] = np.nan
    # Add new TTP90 columns
    df['TTP90 Reached'] = 0
    df['TTP90 Time'] = np.nan
    df['TTP90 Threshold Required (μN)'] = np.nan
    # Add new metrics from v1.0.3
    df['TTMFP (s)'] = np.nan
    df['Time Over 90% (%)'] = np.nan


    for well in df['Well'].unique():
        well_data = df[df['Well'] == well]
        
        for stim_num, (t0, t1, hz) in enumerate(t0_t1_hz_triplets, 1):
            # Round the t0 and t1 values to 0.01s precision
            t0 = round(t0 * 100) / 100
            t1 = round(t1 * 100) / 100
            
            # Find the closest time point to T1 for Stim Peak
            if len(well_data[well_data['Time (seconds)'] >= t1]) == 0:
                # print(f"Warning: No data points found at or after time {t1} for well {well}. Skipping stim {stim_num} ({hz}Hz).")
                continue
                
            stim_peak_index = well_data['Time (seconds)'].sub(t1).abs().idxmin()
            df.loc[stim_peak_index, 'Stim Peak'] = 1
            df.loc[stim_peak_index, 'Stim Number'] = stim_num  # Store stimulation number
            df.loc[stim_peak_index, 'Hz'] = hz                 # Store Hz as separate column
            stim_peak_time = df.loc[stim_peak_index, 'Time (seconds)']
            # Round the peak time to 0.01s precision
            stim_peak_time = round(stim_peak_time * 100) / 100
            
            # Search for maximum in ±100ms window around T1 for kinetics calculations
            expanded_peak_window = well_data[(well_data['Time (seconds)'] >= t1 - 0.1) & 
                                             (well_data['Time (seconds)'] <= t1 + 0.1)]
            
            if len(expanded_peak_window) == 0:
                # print(f"Warning: No data points found in kinetics peak window for well {well}. Using fallback method.")
                # Keep using the 50ms window as before for backward compatibility
                peak_window = well_data[(well_data['Time (seconds)'] >= stim_peak_time - 0.05) & 
                                     (well_data['Time (seconds)'] <= stim_peak_time)]
                
                if len(peak_window) == 0:
                    # print(f"Warning: No data points found in peak window for well {well}. Skipping.")
                    continue
                    
                kinetics_peak_force = peak_window['Active Twitch Force (μN)_Smoothed'].mean()
                df.loc[stim_peak_index, 'Kinetics Peak Force (μN)'] = kinetics_peak_force
                kinetics_peak_time = stim_peak_time
            else:
                # Find the maximum force value and its time for kinetics
                max_idx = expanded_peak_window['Active Twitch Force (μN)_Smoothed'].idxmax()
                kinetics_peak_force = expanded_peak_window.loc[max_idx, 'Active Twitch Force (μN)_Smoothed']
                kinetics_peak_time = expanded_peak_window.loc[max_idx, 'Time (seconds)']
                
                # Store kinetics peak values
                df.loc[stim_peak_index, 'Kinetics Peak Force (μN)'] = kinetics_peak_force
                df.loc[stim_peak_index, 'Kinetics Peak Time (seconds)'] = kinetics_peak_time
                df.loc[stim_peak_index, 'Peak Time Delta (s)'] = kinetics_peak_time - stim_peak_time
                
                # if abs(kinetics_peak_time - stim_peak_time) > 0.03:  # If more than 30ms difference
                #     print(f"Well {well}, Stim {stim_num} ({hz}Hz): Kinetics peak at {kinetics_peak_time:.2f}s (Δ={kinetics_peak_time-stim_peak_time:.2f}s) from T1")
            
            # Calculate baseline using configurable window
            baseline_window = well_data[(well_data['Time (seconds)'] >= t0 - BASELINE_WINDOW) & 
                                       (well_data['Time (seconds)'] < t0)]
            
            if len(baseline_window) == 0:
                # print(f"Warning: No data points found in baseline window for well {well}. Using 0 as baseline.")
                baseline = 0
            else:
                baseline = baseline_window['Active Twitch Force (μN)_Smoothed'].mean()
                
            # Store baseline value in the dataframe
            df.loc[stim_peak_index, 'Baseline (μN)'] = baseline

            # SEPARATE CALCULATION: Max Tetanic Force (T0 to T1+200ms search) - includes late peaks
            tetanic_window = well_data[(well_data['Time (seconds)'] >= t0) & 
                                     (well_data['Time (seconds)'] <= t1 + 0.2)]  # Extended 200ms past T1
            
            if len(tetanic_window) > 0:
                max_tetanic_force_raw = tetanic_window['Active Twitch Force (μN)_Smoothed'].max()
                max_tetanic_bk = max_tetanic_force_raw - baseline  # BK correction applied
                df.loc[stim_peak_index, 'Max Tetanic Force (μN)'] = max_tetanic_force_raw
                df.loc[stim_peak_index, 'Max tetanic-BK (μN)'] = max_tetanic_bk
                # print(f"Well {well}, Stim {stim_num} ({hz}Hz): Max Tetanic = {max_tetanic_force_raw:.1f}μN, Max Tetanic-BK = {max_tetanic_bk:.1f}μN, Kinetics Peak = {kinetics_peak_force:.1f}μN")
            else:
                # print(f"Warning: No data points found in tetanic window (T0-T1+200ms) for well {well}.")
                df.loc[stim_peak_index, 'Max Tetanic Force (μN)'] = kinetics_peak_force  # Fallback
                kinetics_bk = kinetics_peak_force - baseline
                df.loc[stim_peak_index, 'Max tetanic-BK (μN)'] = kinetics_bk  # Fallback
            
            # Calculate BK corrected force using kinetics peak for relaxation calculations
            bk_corrected_force = kinetics_peak_force - baseline
            df.loc[stim_peak_index, 'Kinetics Peak-BK (μN)'] = bk_corrected_force
            
            # === LOGIC GATE: Only calculate kinetics if Hz > 20 ===
            if hz > 20:
                # --- START: New Metrics Calculation (v1.0.3) ---
                # 1. Time to Max Force Peak (TTMFP)
                # Define the search window from T0 for 2.5 seconds
                ttmfp_window = well_data[(well_data['Time (seconds)'] >= t0) &
                                         (well_data['Time (seconds)'] <= t0 + 2.5)]

                if not ttmfp_window.empty:
                    # Find the peak force within this window from the smoothed data
                    max_idx = ttmfp_window['Active Twitch Force (μN)_Smoothed'].idxmax()
                    peak_force_for_metrics = ttmfp_window.loc[max_idx, 'Active Twitch Force (μN)_Smoothed']
                    peak_time_for_metrics = ttmfp_window.loc[max_idx, 'Time (seconds)']

                    # Calculate TTMFP as the time from T0
                    ttmfp = peak_time_for_metrics - t0
                    df.loc[stim_peak_index, 'TTMFP (s)'] = ttmfp

                    # 2. Time Over 90% Force (Fatigue Resistance)
                    # Use the baseline-corrected peak force to set the threshold
                    bk_peak_force = peak_force_for_metrics - baseline
                    threshold_90 = (bk_peak_force * 0.9) + baseline

                    # Analyze the duration within the primary stimulation window (T0 to T1)
                    stim_window = well_data[(well_data['Time (seconds)'] >= t0) &
                                            (well_data['Time (seconds)'] <= t1)]

                    # Count the number of data points above the 90% threshold
                    points_over_90 = stim_window[stim_window['Active Twitch Force (μN)_Smoothed'] >= threshold_90]

                    # Calculate total time (assuming 100Hz data acquisition, i.e., 0.01s per point)
                    time_interval = 0.01 
                    time_over_90_seconds = len(points_over_90) * time_interval
                    
                    # Convert to percentage of the stimulation duration
                    stim_duration = t1 - t0
                    time_over_90_percent = (time_over_90_seconds / stim_duration) * 100 if stim_duration > 0 else 0
                    df.loc[stim_peak_index, 'Time Over 90% (%)'] = time_over_90_percent

                # --- END: New Metrics Calculation ---

                r10_threshold = (bk_corrected_force * 0.9) + baseline
                r50_threshold = (bk_corrected_force * 0.5) + baseline
                r80_threshold = (bk_corrected_force * 0.2) + baseline
                r90_threshold = (bk_corrected_force * 0.1) + baseline
                
                # Store thresholds with full precision
                df.loc[stim_peak_index, 'R10 Threshold Required (μN)'] = r10_threshold
                df.loc[stim_peak_index, 'R50 Threshold Required (μN)'] = r50_threshold
                df.loc[stim_peak_index, 'R80 Threshold Required (μN)'] = r80_threshold
                df.loc[stim_peak_index, 'R90 Threshold Required (μN)'] = r90_threshold
                
                relaxation_window = well_data[well_data['Time (seconds)'] > stim_peak_time]
                
                # ENHANCED RELAXATION DETECTION - ALIGNED WITH TTP METHODS
                for r_name, r_threshold in [('R10', r10_threshold), ('R50', r50_threshold), 
                                            ('R80', r80_threshold), ('R90', r90_threshold)]:
                    consecutive_count = 0
                    r_start_idx = None
                    
                    for idx, row in relaxation_window.iterrows():
                        # Add 102% tolerance for relaxation (allows 2% above threshold for noise)
                        if row['Active Twitch Force (μN)_Smoothed'] <= (r_threshold * 1.02):
                            if consecutive_count == 0:
                                r_start_idx = idx
                            consecutive_count += 1
                            if consecutive_count == 3:  # 3 consecutive points
                                df.loc[r_start_idx, f'{r_name} Reached'] = 1
                                time_diff = df.loc[r_start_idx, 'Time (seconds)'] - stim_peak_time
                                # Round to 0.01s precision
                                time_diff = round(time_diff * 100) / 100
                                df.loc[stim_peak_index, f'{r_name} Time'] = time_diff
                                break
                        else:
                            # Gradual decay instead of reset to 0
                            consecutive_count = max(0, consecutive_count - 1)
                    
                    # Fallback method if no 3 consecutive points found
                    if pd.isna(df.loc[stim_peak_index, f'{r_name} Time']):
                        below_threshold = relaxation_window[
                            relaxation_window['Active Twitch Force (μN)_Smoothed'] <= r_threshold]
                        
                        if not below_threshold.empty:
                            first_idx = below_threshold.index[0]
                            time_diff = below_threshold.loc[first_idx, 'Time (seconds)'] - stim_peak_time
                            time_diff = round(time_diff * 100) / 100
                            df.loc[stim_peak_index, f'{r_name} Time'] = time_diff
                            df.loc[stim_peak_index, f'{r_name} Reached'] = 1
                            # print(f"Well {well}, Stim {stim_num} ({hz}Hz): Using single-point method for {r_name} at time {time_diff}s")
                
                # Calculate TT50P (Time to 50% Peak) - ENHANCED
                tt50p_threshold = r50_threshold  # Same as R50 threshold
                df.loc[stim_peak_index, 'TT50P Threshold Required (μN)'] = tt50p_threshold
                
                # Extended contraction window to allow for slower rises
                max_window_time = min(t0 + 5.0, well_data['Time (seconds)'].max())
                extended_contraction_window = well_data[(well_data['Time (seconds)'] >= t0) & 
                                                        (well_data['Time (seconds)'] <= max_window_time)]
                
                consecutive_count = 0
                tt50p_start_idx = None
                
                for idx, row in extended_contraction_window.iterrows():
                    # Add small tolerance to handle noise
                    if row['Active Twitch Force (μN)_Smoothed'] >= (tt50p_threshold * 0.98):
                        if consecutive_count == 0:
                            tt50p_start_idx = idx
                        consecutive_count += 1
                        if consecutive_count == 3:  # 3 consecutive points
                            df.loc[tt50p_start_idx, 'TT50P Reached'] = 1
                            time_diff = df.loc[tt50p_start_idx, 'Time (seconds)'] - t0
                            # Round to 0.01s precision
                            time_diff = round(time_diff * 100) / 100
                            df.loc[stim_peak_index, 'TT50P Time'] = time_diff
                            break
                    else:
                        # Don't reset consecutive count to 0, just decrement by 1
                        consecutive_count = max(0, consecutive_count - 1)
                
                # Fallback method if no 3 consecutive points found
                if pd.isna(df.loc[stim_peak_index, 'TT50P Time']):
                    above_threshold = extended_contraction_window[
                        extended_contraction_window['Active Twitch Force (μN)_Smoothed'] >= tt50p_threshold]
                    
                    if not above_threshold.empty:
                        first_idx = above_threshold.index[0]
                        time_diff = above_threshold.loc[first_idx, 'Time (seconds)'] - t0
                        time_diff = round(time_diff * 100) / 100
                        df.loc[stim_peak_index, 'TT50P Time'] = time_diff
                        df.loc[stim_peak_index, 'TT50P Reached'] = 1
                        # print(f"Well {well}, Stim {stim_num} ({hz}Hz): Using single-point method for TT50P at time {time_diff}s")
                
                # Calculate TTP90 (Time to 90% Peak)
                ttp90_threshold = (bk_corrected_force * 0.9) + baseline
                df.loc[stim_peak_index, 'TTP90 Threshold Required (μN)'] = ttp90_threshold
                
                # Extended contraction window to allow for slower rises
                max_window_time = min(t0 + 5.0, well_data['Time (seconds)'].max())
                extended_contraction_window = well_data[(well_data['Time (seconds)'] >= t0) & 
                                                        (well_data['Time (seconds)'] <= max_window_time)]
                
                consecutive_count = 0
                ttp90_start_idx = None
                
                for idx, row in extended_contraction_window.iterrows():
                    # Add small tolerance to handle noise
                    if row['Active Twitch Force (μN)_Smoothed'] >= (ttp90_threshold * 0.98):
                        if consecutive_count == 0:
                            ttp90_start_idx = idx
                        consecutive_count += 1
                        if consecutive_count == 3:  # 3 consecutive points
                            df.loc[ttp90_start_idx, 'TTP90 Reached'] = 1
                            time_diff = df.loc[ttp90_start_idx, 'Time (seconds)'] - t0
                            # Round to 0.01s precision
                            time_diff = round(time_diff * 100) / 100
                            df.loc[stim_peak_index, 'TTP90 Time'] = time_diff
                            break
                    else:
                        # Don't reset consecutive count to 0, just decrement by 1
                        consecutive_count = max(0, consecutive_count - 1)
                
                # Fallback method if no 3 consecutive points found
                if pd.isna(df.loc[stim_peak_index, 'TTP90 Time']):
                    above_threshold = extended_contraction_window[
                        extended_contraction_window['Active Twitch Force (μN)_Smoothed'] >= ttp90_threshold]
                    
                    if not above_threshold.empty:
                        first_idx = above_threshold.index[0]
                        time_diff = above_threshold.loc[first_idx, 'Time (seconds)'] - t0
                        time_diff = round(time_diff * 100) / 100
                        df.loc[stim_peak_index, 'TTP90 Time'] = time_diff
                        df.loc[stim_peak_index, 'TTP90 Reached'] = 1
                        # print(f"Well {well}, Stim {stim_num} ({hz}Hz): Using single-point method for TTP90 at time {time_diff}s")

    return df

def process_continuous_waveforms(df, plate_map, plate_name, max_time):
    try:
        relevant_columns = ['Time (seconds)'] + [f'{row}{col} - Active Twitch Force (μN)' 
                                                 for row in 'ABCD' for col in range(1, 7)]
        df_filtered = df[df.columns.intersection(relevant_columns)]
        
        # Filter data to only include times up to max_time
        if 'Time (seconds)' in df_filtered.columns:
            df_filtered = df_filtered[df_filtered['Time (seconds)'] <= max_time]
        
        for col in df_filtered.columns:
            if col != 'Time (seconds)':
                df_filtered = df_filtered[df_filtered[col] != 0]
        
        # Round time values to 0.01s precision (instead of mapping)
        df_filtered['Time (seconds)'] = (df_filtered['Time (seconds)'] * 100).round() / 100
        
        data_columns = [col for col in df_filtered.columns if col != 'Time (seconds)' and not df_filtered[col].isnull().all()]
        
        # Check if any data columns remain
        if not data_columns:
            # print(f"Warning: No valid data columns found after filtering for {plate_name}")
            return pd.DataFrame()  # Return empty dataframe
            
        melted_df = df_filtered.melt(id_vars='Time (seconds)', value_vars=data_columns, var_name='Well', value_name='Active Twitch Force (μN)')
        melted_df['Well'] = melted_df['Well'].str.split(' - ').str[0]
        
        def get_condition(well):
            try:
                condition = plate_map.loc[well[0], well[1]]
                return condition
            except KeyError:
                # print(f"Warning: Well {well} not found in plate map. Setting condition to 'Unknown'.")
                return 'Unknown'
        
        melted_df['Condition'] = melted_df['Well'].apply(get_condition)
        
        all_time_points = set(melted_df['Time (seconds)'])
        for well in melted_df['Well'].unique():
            well_time_points = set(melted_df[melted_df['Well'] == well]['Time (seconds)'])
            all_time_points = all_time_points.intersection(well_time_points)
        
        # Check if we have any common time points
        if not all_time_points:
            # print(f"Warning: No common time points found across wells for {plate_name}")
            return pd.DataFrame()  # Return empty dataframe
            
        melted_df = melted_df[melted_df['Time (seconds)'].isin(all_time_points)]
        
        melted_df['Plate'] = plate_name
        
        # Apply smoothing first
        melted_df = smooth_data(melted_df, 'Active Twitch Force (μN)', window_size=3)
        
        # Normalize from smoothed data instead of raw data
        well_stats = melted_df.groupby('Well')['Active Twitch Force (μN)_Smoothed'].agg(['min', 'max'])
        melted_df = melted_df.merge(well_stats, left_on='Well', right_index=True)
        
        # Avoid division by zero in normalization
        melted_df['range'] = melted_df['max'] - melted_df['min']
        # Keep full precision for normalized values
        melted_df['Normalized Trace'] = np.where(
            melted_df['range'] > 0,
            (melted_df['Active Twitch Force (μN)_Smoothed'] - melted_df['min']) / melted_df['range'],
            0
        )
        
        # Check if 'Normalized Trace' column was created successfully
        if 'Normalized Trace' not in melted_df.columns:
            # print(f"Warning: Failed to create 'Normalized Trace' column for {plate_name}. Creating default.")
            melted_df['Normalized Trace'] = 0
        
        melted_df = melted_df.drop(columns=['min', 'max', 'range'])
        
        # Now proceed with peak detection and other calculations
        melted_df = find_peaks_and_relaxation_times(melted_df, T0_T1_HZ_TRIPLETS)
      
        column_order = ['Time (seconds)', 'Plate', 'Well', 'Condition', 
                      'Active Twitch Force (μN)', 'Active Twitch Force (μN)_Smoothed',
                      'Normalized Trace', 'Stim Peak', 'Stim Number', 'Hz',
                      'Kinetics Peak Force (μN)', 'Max Tetanic Force (μN)', 'Baseline (μN)', 
                      'Kinetics Peak-BK (μN)', 'Max tetanic-BK (μN)',
                      'Kinetics Peak Time (seconds)', 'Peak Time Delta (s)', 
                      'R10 Reached', 'R10 Time', 'R10 Threshold Required (μN)',
                      'R50 Reached', 'R50 Time', 'R50 Threshold Required (μN)',
                      'R80 Reached', 'R80 Time', 'R80 Threshold Required (μN)',
                      'R90 Reached', 'R90 Time', 'R90 Threshold Required (μN)',
                      'TT50P Reached', 'TT50P Time', 'TT50P Threshold Required (μN)',
                      'TTP90 Reached', 'TTP90 Time', 'TTP90 Threshold Required (μN)',
                      'TTMFP (s)', 'Time Over 90% (%)'] # Added new metrics
        
        # Ensure all expected columns exist
        for col in column_order:
            if col not in melted_df.columns:
                melted_df[col] = np.nan
                
        melted_df = melted_df[column_order]
        
        return melted_df
        
    except Exception as e:
        print(f"Error in process_continuous_waveforms for {plate_name}: {e}")
        return pd.DataFrame()  # Return empty dataframe on error

def calculate_average_trace(data_list):
    if not data_list:
        return pd.DataFrame(columns=['Time (seconds)', 'Normalized Trace', 'Standard Deviation', 
                                    'Force (μN)', 'Force SD', 'Smoothed Force (μN)', 'Smoothed Force SD'])
    
    try:
        combined_data = pd.concat(data_list, ignore_index=True)
        # Round time values to ensure consistent grouping
        combined_data['Time (seconds)'] = (combined_data['Time (seconds)'] * 100).round() / 100
        
        # Start with time as the base column
        avg_trace = pd.DataFrame({'Time (seconds)': sorted(combined_data['Time (seconds)'].unique())})
        
        # Create mapping functions that handle missing data gracefully
        def safe_map(df, time_col, value_col, map_to_df):
            """Safely map values from one df to another with error handling"""
            if value_col not in df.columns:
                # print(f"Warning: Column '{value_col}' not found in data. Skipping.")
                return np.nan
            
            try:
                # Create the groupby object
                grouped = df.groupby(time_col)
                
                # Calculate mean and reset index to get a mappable dataframe
                values_mean = grouped[value_col].mean().reset_index()
                values_std = grouped[value_col].std().reset_index()
                
                # Create mapping dictionaries
                mean_map = dict(zip(values_mean[time_col], values_mean[value_col]))
                std_map = dict(zip(values_std[time_col], values_std[value_col]))
                
                # Return the mapped values
                return map_to_df[time_col].map(mean_map), map_to_df[time_col].map(std_map)
            except Exception as e:
                print(f"Error mapping {value_col}: {e}")
                return np.nan, np.nan
        
        # Apply the mapping function to each column of interest
        if 'Normalized Trace' in combined_data.columns:
            avg_trace['Normalized Trace'], avg_trace['Standard Deviation'] = safe_map(
                combined_data, 'Time (seconds)', 'Normalized Trace', avg_trace)
        else:
            # print("Warning: 'Normalized Trace' column not found in data")
            avg_trace['Normalized Trace'] = np.nan
            avg_trace['Standard Deviation'] = np.nan
        
        if 'Active Twitch Force (μN)' in combined_data.columns:
            avg_trace['Force (μN)'], avg_trace['Force SD'] = safe_map(
                combined_data, 'Time (seconds)', 'Active Twitch Force (μN)', avg_trace)
        else:
            avg_trace['Force (μN)'] = np.nan
            avg_trace['Force SD'] = np.nan
        
        if 'Active Twitch Force (μN)_Smoothed' in combined_data.columns:
            avg_trace['Smoothed Force (μN)'], avg_trace['Smoothed Force SD'] = safe_map(
                combined_data, 'Time (seconds)', 'Active Twitch Force (μN)_Smoothed', avg_trace)
        else:
            avg_trace['Smoothed Force (μN)'] = np.nan
            avg_trace['Smoothed Force SD'] = np.nan
        
        return avg_trace
    
    except Exception as e:
        print(f"Error in calculate_average_trace: {e}")
        # Return an empty dataframe with expected columns
        return pd.DataFrame(columns=['Time (seconds)', 'Normalized Trace', 'Standard Deviation', 
                                    'Force (μN)', 'Force SD', 'Smoothed Force (μN)', 'Smoothed Force SD'])

def generate_calculation_reference(df, t0_t1_hz_triplets, output_dir, day, protocol):
    """Generate a detailed calculation reference document with example calculations"""
    
    # Find any well with peak data
    example_well = None
    example_data = None
    
    if 'Stim Peak' in df.columns:
        # Get all rows with peaks
        peaks = df[df['Stim Peak'] == 1]
        
        if not peaks.empty:
            # Take the first peak from the first well with peaks
            first_peak = peaks.iloc[0]
            example_well = first_peak['Well']
            example_data = first_peak
            print(f"Using well {example_well} for calculation reference")
    
    if example_data is None:
        print(f"No wells with peak data found for calculation reference")
        return
    
    # Get timestamps for reference - round to 0.01s precision
    t0, t1, hz = t0_t1_hz_triplets[0]  # First stimulation
    t0 = round(t0 * 100) / 100
    t1 = round(t1 * 100) / 100
    peak_time = round(example_data['Time (seconds)'] * 100) / 100
    
    # Get condition and stimulation info
    condition = example_data['Condition']
    stim_number = example_data.get('Stim Number', 'N/A')
    stim_hz = example_data.get('Hz', 'N/A')
    
    # Create reference document with error handling for missing fields
    reference = [
        f"# Force vs Frequency (FvF) Analysis - Calculation Reference ({day}_{protocol})",
        "",
        f"This document explains the calculation methods used in the FvF waveform analysis script, with example calculations from Well {example_well}.",
        "",
        f"## Example Calculation for Well {example_well} ({condition})",
        f"### Stimulation: #{stim_number} ({stim_hz}Hz)",
        "",
        f"### FvF Protocol Parameters",
        f"- T0 (stimulation start): {t0:.2f} seconds",
        f"- T1 (stimulation end): {t1:.2f} seconds", 
        f"- Stimulation frequency: {stim_hz}Hz",
        f"- Detected Peak Time: {peak_time:.2f} seconds"
    ]
    
    # Add fields that might be missing with error handling
    if 'Kinetics Peak Time (seconds)' in example_data and not pd.isna(example_data['Kinetics Peak Time (seconds)']):
        reference.append(f"- Kinetics Peak Time: {example_data['Kinetics Peak Time (seconds)']:.2f} seconds (Δ={example_data['Peak Time Delta (s)']:.2f}s)")
    
    reference.extend([
        "",
        f"### Force Measurements"
    ])
    
    if 'Baseline (μN)' in example_data and not pd.isna(example_data['Baseline (μN)']):
        reference.append(f"- Baseline Force: {example_data['Baseline (μN)']} μN")
    
    if 'Kinetics Peak Force (μN)' in example_data and not pd.isna(example_data['Kinetics Peak Force (μN)']):
        reference.append(f"- Kinetics Peak Force: {example_data['Kinetics Peak Force (μN)']} μN")
    
    if 'Max Tetanic Force (μN)' in example_data and not pd.isna(example_data['Max Tetanic Force (μN)']):
        reference.append(f"- Max Tetanic Force (raw): {example_data['Max Tetanic Force (μN)']} μN")
    
    if 'Max tetanic-BK (μN)' in example_data and not pd.isna(example_data['Max tetanic-BK (μN)']):
        reference.append(f"- Max Tetanic-BK Force: {example_data['Max tetanic-BK (μN)']} μN")
    
    if 'Kinetics Peak-BK (μN)' in example_data and not pd.isna(example_data['Kinetics Peak-BK (μN)']):
        reference.append(f"- Kinetics Peak-BK Force: {example_data['Kinetics Peak-BK (μN)']} μN")
    
    reference.extend([
        "",
        f"### Threshold Calculations"
    ])
    
    threshold_fields = [
        ('R10 Threshold Required (μN)', 'R10 Threshold'),
        ('R50 Threshold Required (μN)', 'R50 Threshold'),
        ('R80 Threshold Required (μN)', 'R80 Threshold'),
        ('R90 Threshold Required (μN)', 'R90 Threshold'),
        ('TTP90 Threshold Required (μN)', 'TTP90 Threshold')
    ]
    
    for field, label in threshold_fields:
        if field in example_data and not pd.isna(example_data[field]):
            reference.append(f"- {label}: {example_data[field]} μN")
    
    reference.extend([
        "",
        f"### Time Measurements"
    ])
    
    time_fields = [
        ('R10 Time', 'R10 Time'),
        ('R50 Time', 'R50 Time'),
        ('R80 Time', 'R80 Time'),
        ('R90 Time', 'R90 Time'),
        ('TT50P Time', 'TT50P Time'),
        ('TTP90 Time', 'TTP90 Time'),
        ('TTMFP (s)', 'TTMFP'), # Added new metric
        ('Time Over 90% (%)', 'Time Over 90%') # Added new metric
    ]
    
    for field, label in time_fields:
        if field in example_data and not pd.isna(example_data[field]):
            reference.append(f"- {label}: {example_data[field]:.2f} {'seconds' if 'Time' in field else '%'}")

    # Add the rest of the documentation
    reference.extend([
        "",
        "## Force vs Frequency (FvF) Protocol",
        "",
        "### Stimulation Schedule:",
        "The FvF protocol consists of 14 stimulations with increasing frequencies:",
        ""
    ])
    
    # Add the stimulation schedule
    for i, hz in enumerate(FVF_PROTOCOL_HZ, 1):
        reference.append(f"  Stim {i:2d}: {hz:3d}Hz")
    
    reference.extend([
        "",
        "### Timing Parameters:",
        f"- Initial delay: {INITIAL_DELAY:.1f} seconds",
        f"- Each stimulation duration: {STIM_DURATION:.1f} seconds", 
        f"- Inter-stimulation delay: {INTER_STIM_DELAY:.1f} seconds",
        f"- Baseline window: {BASELINE_WINDOW:.1f} seconds before T0",
        "",
        "## General Calculation Methods",
        "",
        "1. **Baseline Force**: Average force over baseline window before each T0",
        "2. **Kinetics Peak Force**: Maximum force value in ±100ms window around T1 (for kinetics calculations)",
        "3. **Max Tetanic Force**: Maximum force in T0 to T1+200ms window (captures late peaks after stimulation)",
        "4. **Max Tetanic-BK**: Max Tetanic Force MINUS baseline (BK corrected for F-F analysis)",
        "5. **Kinetics Peak-BK**: Kinetics Peak Force - Baseline (used for relaxation thresholds)",
        "6. **FF Normalized (%)**: Each stimulation normalized as % of maximum Max Tetanic-BK within that well across all Hz",
        "7. **Thresholds** (based on kinetics peak):",
        "    - R10: Baseline + (Kinetics Peak-BK × 0.9)",
        "    - R50: Baseline + (Kinetics Peak-BK × 0.5)",
        "    - R80: Baseline + (Kinetics Peak-BK × 0.2)",
        "    - R90: Baseline + (Kinetics Peak-BK × 0.1)",
        "8. **Relaxation Times**: Time from T1 until force drops below threshold",
        "9. **TT50P**: Time from T0 until force rises above 50% of peak force",
        "10. **TTP90**: Time from T0 until force rises above 90% of peak force",
        "11. **TTMFP (s)**: Time from T0 to the absolute maximum force within a 2.5s window.",
        "12. **Time Over 90% (%)**: Percentage of stimulation time (T0-T1) spent above 90% of the max force found for TTMFP. Measures fatigue resistance.",
        "",
        "## Key Changes in v1.0.1",
        "",
        "### Max Tetanic Force Search Window Extended:",
        "- **Old**: T0 to T1 search window (missed late peaks after stimulation)",
        "- **New**: T0 to T1+200ms search window (captures late peaks that occur after stimulation ends)",
        "- **Benefit**: Ensures Max Tetanic Force >= Kinetics Peak Force (logical consistency)",
        "",
        "### Dual Max Tetanic Force Columns:",
        "- **Max Tetanic Force (μN)**: Raw maximum force in T0 to T1+200ms window",
        "- **Max tetanic-BK (μN)**: Same as above but baseline corrected (Max Tetanic - Baseline)",
        "- **Column Order**: Kinetics Peak → Max Tetanic (raw) → Baseline → Kinetics Peak-BK → Max tetanic-BK",
        "",
        "### Force-Frequency Analysis Benefits:",
        "- **Extended search window** captures all peak types including late peaks",
        "- **Eliminates baseline drift effects** across different Hz stimulations",
        "- **More accurate normalization** for force-frequency curves", 
        "- **Better comparison** between conditions with different baseline forces",
        "- **Logical consistency**: Max Tetanic always >= Kinetics Peak",
        "- **Consistent naming**: All BK corrected columns follow same pattern",
        "",
        "## Enhanced Detection Methods (FvF v1.0)",
        "",
        "### Relaxation Kinetics Detection",
        "- **Consecutive Points**: 3 consecutive points below threshold",
        "- **Tolerance**: 102% tolerance (allows 2% above threshold for noise)",
        "- **Gradual Decay**: Count decrements by 1 instead of reset to 0",
        "- **Fallback**: Single-point detection if consecutive method fails",
        "",
        "### Contraction Kinetics Detection", 
        "- **Consecutive Points**: 3 consecutive points above threshold",
        "- **Tolerance**: 98% tolerance (allows 2% below threshold for noise)",
        "- **Gradual Decay**: Count decrements by 1 instead of reset to 0",
        "- **Fallback**: Single-point detection if consecutive method fails",
        "",
        "## Data Structure for Force-Frequency Analysis",
        "",
        "### Key Columns for Analysis:",
        "- **Stim Number**: Sequential stimulation number (1-14)",
        "- **Hz**: Stimulation frequency (separate filterable column)",
        "- **Max Tetanic Force (μN)**: Raw maximum force in T0 to T1+200ms window",
        "- **Max tetanic-BK (μN)**: BK corrected maximum tetanic force (for absolute force analysis)",
        "- **FF Normalized (%)**: Force as % of maximum within each well (ideal for F-F curves)",
        "- **All kinetics metrics**: R10/50/80/90 Time, TT50P Time, TTP90 Time",
        "",
        "### Recommended Analysis Approach:",
        "1. **Force-Frequency Curves**: Plot FF Normalized (%) vs Hz for clean F-F relationships",
        "2. **Absolute Force Curves**: Plot Max tetanic-BK vs Hz for baseline-corrected absolute values",
        "3. **Raw Force Curves**: Plot Max Tetanic Force vs Hz for raw absolute values",
        "4. **Kinetics vs Frequency**: Plot relaxation times vs Hz", 
        "5. **Condition Comparisons**: Group by Condition and Hz for statistics",
        "6. **Peak Detection Window**: ±100ms around T1 (kinetics), T0 to T1+200ms (max tetanic)",
        "",
        "## Processing and Normalization",
        "",
        "1. **Data Smoothing**: A centered rolling mean with window size 3 is applied to raw force",
        "2. **Normalization**: The smoothed force is normalized to a 0-1 scale for each well",
        "3. **Peak Detection**: All metrics are calculated from the smoothed force data",
        "4. **Representative Traces**: Include both raw and smoothed force averages across wells",
        "",
        "## Relaxation Time Interpretation",
        "",
        "- **R10 Time**: Very early relaxation - typically 0.01-0.03s for healthy tissue",
        "- **R50 Time**: Half-relaxation time - typically 0.07-0.15s",
        "- **R80 Time**: Late relaxation - typically 0.20-0.40s", 
        "- **R90 Time**: Near-complete relaxation - typically 0.40-0.70s",
        "",
        "## Contraction Time Interpretation",
        "",
        "- **TT50P**: Time to reach 50% of peak force - characterizes early phase contraction",
        "- **TTP90**: Time to reach 90% of peak force - characterizes late phase contraction",
        "",
        "## Force-Frequency Relationship",
        "",
        "### Expected Patterns:",
        "- **Peak Force**: Generally increases with frequency up to optimal Hz, then plateaus/decreases",
        "- **Kinetics**: May show frequency-dependent changes in contraction and relaxation rates",
        "- **Individual Variation**: Different conditions may show distinct F-F relationships",
        "",
        "### Analysis Tips:",
        "- **Hz Column**: Use for filtering and grouping in analysis software",
        "- **Multiple Replicates**: Each well provides one data point per frequency",
        "- **Statistical Analysis**: Consider both within-well and between-well variation",
        "- **BK Correction**: All force values now baseline corrected for better comparisons",
        "- **Consistent Naming**: Kinetics Peak-BK and Max tetanic-BK follow same pattern",
        "",
        "## Note on Timing",
        "",
        "All time values are reported with centisecond precision (2 decimal places).",
        "This matches the data acquisition rate of 100Hz (0.01s intervals).",
        "Force values and normalized trace values are stored with full precision.",
        "",
        "## Representative Traces",
        "",
        "The Representative Traces sheet includes:",
        "",
        "1. **Normalized Trace**: Smoothed force normalized to 0-1 scale for shape comparison",
        "2. **Raw Force**: Absolute force values in μN",
        "3. **Smoothed Force**: Smoothed force values used for all kinetics calculations",
        "4. **Standard Deviations**: Variability measures for all three data types",
        "",
        "Note: Individual twitch normalization can be derived from Representative Traces if needed."
    ])
    
    # Replace Greek mu with 'u' in unit symbols to avoid encoding issues
    reference = [line.replace('μN', 'uN') for line in reference]
    
    # Write to file with explicit UTF-8 encoding
    reference_path = os.path.join(output_dir, f"{day}_{protocol}_FvF_calculation_reference.md")
    with open(reference_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(reference))
    
    print(f"FvF calculation reference written to {reference_path}")

def create_relaxation_summary(data_list):
    if not data_list:
        return pd.DataFrame()  # Return empty dataframe if no data
    
    try:
        combined_data = pd.concat(data_list, ignore_index=True)
        
        # Round only time values to 0.01s precision
        time_cols = [col for col in combined_data.columns if 'Time' in col and col != 'Relative Time (ms)']
        for col in time_cols:
            if col in combined_data.columns:
                combined_data[col] = (combined_data[col] * 100).round() / 100
        
        summary = combined_data[combined_data['Stim Peak'] == 1].copy()
        
        if summary.empty:
            return pd.DataFrame()  # Return empty dataframe if no peaks found
        
        # Define columns to extract, checking if they exist
        column_sets = [
            ['Plate', 'Well', 'Condition', 'Time (seconds)', 'Stim Number', 'Hz'],
            ['Kinetics Peak Force (μN)', 'Max Tetanic Force (μN)', 'Baseline (μN)', 'Kinetics Peak-BK (μN)', 'Max tetanic-BK (μN)'],
            ['Kinetics Peak Time (seconds)', 'Peak Time Delta (s)'],
            ['R10 Time', 'R10 Threshold Required (μN)'],
            ['R50 Time', 'R50 Threshold Required (μN)'],
            ['R80 Time', 'R80 Threshold Required (μN)'],
            ['R90 Time', 'R90 Threshold Required (μN)'],
            ['TT50P Time', 'TT50P Threshold Required (μN)'],
            ['TTP90 Time', 'TTP90 Threshold Required (μN)'],
            ['TTMFP (s)', 'Time Over 90% (%)'] # Added new metrics
        ]
        
        columns_to_use = []
        for col_set in column_sets:
            columns_to_use.extend([col for col in col_set if col in summary.columns])
        
        summary = summary[columns_to_use]
        
        if 'Time (seconds)' in summary.columns:
            summary.rename(columns={'Time (seconds)': 'Stim Peak Time'}, inplace=True)
        
        # Calculate FF Normalized (%) - NOW USING BK CORRECTED Max tetanic-BK values
        if 'Max tetanic-BK (μN)' in summary.columns:
            # For each well, find the maximum BK-corrected tetanic force across all Hz stimulations
            well_max_forces = summary.groupby(['Plate', 'Well', 'Condition'])['Max tetanic-BK (μN)'].transform('max')
            
            # Calculate FF normalized percentage using BK corrected values
            summary['FF Normalized (%)'] = (summary['Max tetanic-BK (μN)'] / well_max_forces) * 100
            summary['FF Normalized (%)'] = summary['FF Normalized (%)'].round(1)  # Round to 1 decimal place
        
        # Determine columns for reordering - including only those that exist
        # REMOVED: All "Well Avg" columns to reduce confusion for collaborators
        all_possible_columns = [
            'Plate', 'Well', 'Condition', 'Stim Peak Time', 'Stim Number', 'Hz',
            'Kinetics Peak Force (μN)', 'Max Tetanic Force (μN)', 'Baseline (μN)', 
            'Kinetics Peak-BK (μN)', 'Max tetanic-BK (μN)', 'FF Normalized (%)',
            'Kinetics Peak Time (seconds)', 'Peak Time Delta (s)',
            'R10 Time', 'R10 Threshold Required (μN)',
            'R50 Time', 'R50 Threshold Required (μN)',
            'R80 Time', 'R80 Threshold Required (μN)',
            'R90 Time', 'R90 Threshold Required (μN)',
            'TT50P Time', 'TT50P Threshold Required (μN)',
            'TTP90 Time', 'TTP90 Threshold Required (μN)',
            'TTMFP (s)', 'Time Over 90% (%)' # Added new metrics
        ]
        
        # Filter to only include columns that exist in the summary
        column_order = [col for col in all_possible_columns if col in summary.columns]
        
        # Reorder columns
        summary = summary[column_order]
        
        return summary
        
    except Exception as e:
        print(f"Error creating relaxation summary: {e}")
        return pd.DataFrame()  # Return empty dataframe on error

def process_and_combine(input_files, plate_map, output_dir, max_recording_time):
    processed_data = {}
    average_traces = {}

    for input_file in tqdm(input_files, desc="Processing files"):
        try:
            # Detecting file type and reading content
            if input_file.lower().endswith('.xlsx'):
                xls = pd.ExcelFile(input_file)
                if 'continuous-waveforms' not in xls.sheet_names:
                    # print(f"'continuous-waveforms' sheet not found in {input_file}. Skipping.")
                    continue
                continuous_waveforms_df = pd.read_excel(xls, 'continuous-waveforms')
            
            elif input_file.lower().endswith('.csv'):
                # Read CSV and manually reconstruct the DataFrame to match Excel format
                try:
                    raw_df = pd.read_csv(input_file)
                    
                    # 1. Handle Time column: CSV often has 'Time (s)', we need 'Time (seconds)'
                    if 'Time (s)' in raw_df.columns:
                        raw_df.rename(columns={'Time (s)': 'Time (seconds)'}, inplace=True)
                    
                    # 2. Create a new DataFrame with the specific column names the script expects:
                    # "A1 - Active Twitch Force (μN)", "B1 - Active Twitch Force (μN)", etc.
                    continuous_waveforms_df = pd.DataFrame()
                    if 'Time (seconds)' in raw_df.columns:
                        continuous_waveforms_df['Time (seconds)'] = raw_df['Time (seconds)']
                    
                    # Find columns that look like well IDs (A1, B2, etc.)
                    for col in raw_df.columns:
                        if col in ['Time (seconds)', 'Unnamed: 0']: 
                            continue
                        # Assuming columns like A1, B1... map them to the expected long format
                        new_col_name = f"{col} - Active Twitch Force (μN)"
                        continuous_waveforms_df[new_col_name] = raw_df[col]
                        
                except Exception as e:
                    print(f"Error processing CSV {input_file}: {e}")
                    continue
            else:
                print(f"Unsupported file format: {input_file}")
                continue

            
            # Round time values to 0.01s precision right after loading
            if 'Time (seconds)' in continuous_waveforms_df.columns:
                continuous_waveforms_df['Time (seconds)'] = (continuous_waveforms_df['Time (seconds)'] * 100).round() / 100
                
            # print(f"\nProcessing file: {input_file}")
            # print("Columns found (first 5):", continuous_waveforms_df.columns.tolist()[:5])

        except Exception as e:
            print(f"Error reading input file {input_file}: {e}")
            continue

        filename = os.path.basename(input_file)
        file_info = parse_filename(filename)
        if not file_info:
            print(f"Could not parse filename: {filename}")
            continue
            
        day, protocol, plate_name = file_info

        continuous_df = process_continuous_waveforms(continuous_waveforms_df, plate_map, plate_name, max_recording_time)
        
        if continuous_df.empty:
            # print(f"No data processed for {filename}. Skipping.")
            continue

        if (day, protocol) not in processed_data:
            processed_data[(day, protocol)] = {}
        
        for condition in continuous_df['Condition'].unique():
            if condition not in processed_data[(day, protocol)]:
                processed_data[(day, protocol)][condition] = []
            condition_df = continuous_df[continuous_df['Condition'] == condition]
            processed_data[(day, protocol)][condition].append(condition_df)

    # Calculate average traces for each day, protocol, and condition
    for (day, protocol), conditions in processed_data.items():
        average_traces[(day, protocol)] = {}
        for condition, data_list in conditions.items():
            average_traces[(day, protocol)][condition] = calculate_average_trace(data_list)

        # Create output file with FF naming
        output_filename = f"{day}_FF_waveformkinetics.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        print(f"Creating workbook for {day}_FF")
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet
        
        all_relaxation_data = []

        for condition, data_list in conditions.items():
            if not data_list:
                # print(f"No data for condition {condition}. Skipping.")
                continue
                
            try:
                combined_continuous = pd.concat(data_list, ignore_index=True)
                
                # Round time values only to 0.01s precision
                if 'Time (seconds)' in combined_continuous.columns:
                    combined_continuous['Time (seconds)'] = (combined_continuous['Time (seconds)'] * 100).round() / 100
                
                # Filter data to just include ±1 second around stim peaks
                filtered_data = []
                for well in combined_continuous['Well'].unique():
                    well_data = combined_continuous[combined_continuous['Well'] == well]
                    if 'Stim Peak' in well_data.columns:
                        stim_peak_times = well_data[well_data['Stim Peak'] == 1]['Time (seconds)']
                        
                        for peak_time in stim_peak_times:
                            # Round peak time to 0.01s precision
                            peak_time = round(peak_time * 100) / 100
                            
                            # Extract ±1 second window around each peak
                            peak_window = well_data[(well_data['Time (seconds)'] >= peak_time - 1) & 
                                                  (well_data['Time (seconds)'] <= peak_time + 1)]
                            if not peak_window.empty:
                                filtered_data.append(peak_window)
                
                # Use filtered data for condition sheet if we have peaks
                if filtered_data:
                    filtered_combined = pd.concat(filtered_data, ignore_index=True)
                    sheet_name = f"{condition}"
                    # print(f"Adding filtered sheet for condition: {condition}")
                    df_to_excel_sheet(wb, sheet_name, filtered_combined)
                else:
                    # Otherwise use the full data
                    sheet_name = f"{condition}"
                    # print(f"Adding sheet for condition: {condition} (no peaks found, using full data)")
                    df_to_excel_sheet(wb, sheet_name, combined_continuous)
                    
                all_relaxation_data.extend(data_list)
                    
            except Exception as e:
                print(f"Error processing condition '{condition}': {e}")
                continue

        # Add the Representative Traces sheet (keep full data)
        if conditions:
            rep_traces_list = []
            for cond in conditions.keys():
                if cond in average_traces[(day, protocol)]:
                    avg_trace = average_traces[(day, protocol)][cond]
                    if not avg_trace.empty:
                        avg_trace['Condition'] = cond
                        rep_traces_list.append(avg_trace)
            
            if rep_traces_list:
                try:
                    rep_traces_df = pd.concat(rep_traces_list, ignore_index=True)
                    df_to_excel_sheet(wb, "Representative Traces", rep_traces_df)
                    print("Added Representative Traces sheet")
                except Exception as e:
                    print(f"Error creating Representative Traces sheet: {e}")

        # Add the Relaxation Summary sheet
        if all_relaxation_data:
            try:
                relaxation_summary = create_relaxation_summary(all_relaxation_data)
                if not relaxation_summary.empty:
                    df_to_excel_sheet(wb, "Relaxation Summary", relaxation_summary)
                    print("Added Relaxation Summary sheet")
            except Exception as e:
                print(f"Error creating Relaxation Summary sheet: {e}")

        print(f"Saving workbook to {output_path}")
        wb.save(output_path)
        print(f"FvF continuous waveform data, representative traces, and relaxation summary for {day} saved to {output_path}")
        
        # Generate calculation reference document for this day/protocol
        if 'combined_continuous' in locals() and combined_continuous is not None:
            try:
                # Print available wells for debugging
                available_wells = combined_continuous['Well'].unique()
                print(f"Available wells: {', '.join(available_wells)}")
                
                # Check if any wells have stim peaks
                if 'Stim Peak' in combined_continuous.columns:
                    peak_wells = combined_continuous[combined_continuous['Stim Peak'] == 1]['Well'].unique()
                    if len(peak_wells) > 0:
                        print(f"Wells with detected peaks: {', '.join(peak_wells)}")
                        generate_calculation_reference(combined_continuous, T0_T1_HZ_TRIPLETS, output_dir, day, protocol)
                        print(f"Generated FvF calculation reference document for {day}_{protocol}")
                    else:
                        print(f"No wells with stim peaks detected for {day}_{protocol}")
                else:
                    print(f"No 'Stim Peak' column found in data for {day}_{protocol}")
            except Exception as e:
                print(f"Error generating calculation reference: {e}")
        else:
            print(f"No data available for {day}_{protocol}")

def main():
    try:
        print("=" * 60)
        print("Force vs Frequency (FvF) Waveform Analysis Script v1.1.0")
        print("=" * 60)
        
        # Parse FvF protocol using the global MAX_RECORDING_TIME
        parse_fvf_protocol(MAX_RECORDING_TIME)
        
        # File selection and processing
        input_files = select_files("Select FF input Excel or CSV files", [
        ("Data files", "*.xlsx *.csv"),  # <--- Default: Shows both
        ("Excel files", "*.xlsx"), 
        ("CSV files", "*.csv"), 
        ("All files", "*.*")
    ])
        if not input_files:
            print("No input files selected. Exiting.")
            return

        plate_map_file = select_files("Select plate map CSV file", [("CSV files", "*.csv")])
        if not plate_map_file:
            print("No plate map file selected. Exiting.")
            return

        output_dir = select_directory("Select output directory for processed files")
        if not output_dir:
            print("No output directory selected. Exiting.")
            return

        try:
            plate_map = pd.read_csv(plate_map_file[0], index_col=0, encoding='utf-8')
            plate_map.index = plate_map.index.astype(str)
            plate_map.columns = plate_map.columns.astype(str)
        except Exception as e:
            print(f"Error loading plate map: {e}")
            return

        process_and_combine(input_files, plate_map, output_dir, MAX_RECORDING_TIME)
        
        print("\n" + "=" * 60)
        print("FvF Analysis Complete!")
        print("=" * 60)
    
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()